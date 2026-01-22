from django.shortcuts import render
from impostazioni.models import RegistroImpreseParametri
from impostazioni.models import UtentiParametri

from .utils import salva_log
from .utils import converti_data

from .verifica_cf import verifica_cf_azienda

##from datetime import datetime, date
import datetime
import uuid
import xmltodict

import jwt
import subprocess
import requests
import json
import io
import csv
###import re
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment
from openpyxl.utils import get_column_letter

from django.http import HttpResponse


def inipec_export_excel(request):
    data = request.session.get("multi_data", [])  # Oppure recuperalo come preferisci
    wb = Workbook()
    ws = wb.active
    ws.append(["#", "CF", "Verifica", "Domicilio digitale"])

    # Dizionario per la larghezza massima di ogni colonna
    max_lengths = [len(cell.value) for cell in ws[1]]  # Larghezze iniziali dall'intestazione

    if not data:
        return HttpResponse("Nessun dato disponibile per l'esportazione", status=400)

    for i, row in enumerate(data, 1):
        split_row = row.split()
        stato = split_row[1]
        color = "FFFFFF"  # default bianco

        if stato == "1" and "@" in split_row[2]:
            color = "C6EFCE"  # verde chiaro
        elif stato == "2":
            color = "FFEB9C"  # giallo
        else:
            color = "FFC7CE"  # rosso

        fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

        ws.append([i, split_row[0], split_row[1], " ".join(split_row[2:])])
        for col in ws.iter_rows(min_row=ws.max_row, max_row=ws.max_row):
            for i, cell in enumerate(col):
                cell.fill = fill
                value_length = len(str(cell.value)) if cell.value else 0
                if len(max_lengths) <= i:
                    max_lengths.append(value_length)
                else:
                    max_lengths[i] = max(max_lengths[i], value_length)

    # Imposta larghezza colonne
    for i, width in enumerate(max_lengths, 1):
        col_letter = get_column_letter(i)
        ws.column_dimensions[col_letter].width = width

    # Scrive l'Excel su un buffer
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=EsitoDomiciliDigitaliAziende.xlsx'
    return response



def inipec_singola(request):
    if request.user.id:
        utente_sessione = UtentiParametri.objects.get(id=request.user.id)
        utente_abilitato = utente_sessione.inipec_singolo
        if request.method == 'POST':
            data = []
            cf = request.POST.get('input_CF')
            data.append(cf)
            bearer = registro_imprese_get_bearer()
            correttezza_cf_azienda = verifica_cf_azienda(cf)
            if correttezza_cf_azienda == 1:
                elenco_dati_registro =  registro_imprese_verifica_utente(cf, bearer)
                if elenco_dati_registro:
                    dd = elenco_dati_registro.get('blocchi_impresa').get('dati_identificativi').get('indirizzo_posta_certificata')
                else:
                    dd="Ditta non presente nel Registro Imprese"
                data.append(dd.lower())
                data = converti_data(data)
            else:
                data.append("Codice fiscale non corretto")
            salva_log(request.user,"Verifica INI-PEC singolo", "Verificato domicilio impresa " + cf )
            return render(request, 'inipec_singola.html', {'data': data, 'utente_abilitato': utente_abilitato })
    else:
        utente_abilitato = False
    return render(request, 'inipec_singola.html', { 'utente_abilitato': utente_abilitato })


def inipec_massiva(request):
    if request.user.id:
        utente_sessione = UtentiParametri.objects.get(id=request.user.id)
        utente_abilitato = utente_sessione.inipec_massivo
        if request.method == 'POST':
            csv_file = request.FILES['cf_csv']
            data = []
            contatore = 0
            bearer = registro_imprese_get_bearer()
            if csv_file.name.endswith('.csv'):
                csv_file_text = io.TextIOWrapper(csv_file.file, encoding='utf-8')
                csv_reader = csv.reader(csv_file_text)
                for row in csv_reader:
                    if row[0]:
                        if len(row[0]) in [8, 9, 10]:
                            row[0] = row[0].zfill(11)
                        correttezza_cf_azienda = verifica_cf_azienda(row[0].strip().upper())
                        if correttezza_cf_azienda == 1:
                            elenco_dati_registro =  registro_imprese_verifica_utente(row[0].strip().upper(), bearer)
                            if elenco_dati_registro:
                                dd = elenco_dati_registro.get('blocchi_impresa').get('dati_identificativi').get('indirizzo_posta_certificata')
                            else:
                                dd="Ditta non presente nel Registro Imprese"
                            data.append(row[0].strip().upper() + " " + str(correttezza_cf_azienda) + " " + dd.lower())
                        else:
                            data.append(row[0].strip().upper() + " " + str(correttezza_cf_azienda) + " Codice fiscale non corretto")
                        contatore += 1
                salva_log(request.user,"Verifica INI-PEC massivo", "Verificati n. " + str(contatore) + " CF")
                request.session["multi_data"] = data  # <--- Salva i dati nella sessione
                return render(request, 'inipec_massiva.html', {'data': data, 'utente_abilitato': utente_abilitato })
            elif csv_file.name.endswith('.xlsx') or csv_file.name.endswith('.XLSX') or csv_file.name.endswith('.Xlsx'):
                wb = openpyxl.load_workbook(csv_file)
                sheet = wb.active
                for row in sheet.iter_rows(min_row=1, values_only=True):
                    if row[0]:
                        row = list(row)
                        if not isinstance(row[0], str):
                            row[0] = str(row[0])
                        if len(row[0]) in [8, 9, 10]:
                            row[0] = row[0].zfill(11)
                        correttezza_cf_azienda = verifica_cf_azienda(row[0].strip().upper())
                        if correttezza_cf_azienda == 1:
                            elenco_dati_registro = registro_imprese_verifica_utente(row[0].strip().upper(), bearer)
                            if elenco_dati_registro:
                                dd = elenco_dati_registro.get('blocchi_impresa').get('dati_identificativi').get('indirizzo_posta_certificata')
                            else:
                                dd="Ditta non presente nel Registro Imprese"
                            data.append(row[0].strip().upper() + " " + str(correttezza_cf_azienda) + " " + dd.lower())
                        else:
                            data.append(row[0].strip().upper() + " " + str(correttezza_cf_azienda) + " Codice fiscale non corretto")
                        contatore += 1

                salva_log(request.user,"Verifica INI-PEC massivo", "Verificati n. " + str(contatore) + " CF")
                request.session["multi_data"] = data  # <--- Salva i dati nella sessione
                return render(request, 'inipec_massiva.html', {'data': data, 'utente_abilitato': utente_abilitato })
            else:
                salva_log(request.user,"Verifica INI-PEC massivo", "Errore caricamento file excel-CSV")
                return render(request, 'inipec_massiva.html', {'error': 'Il file non è un XLSX o CSV', 'utente_abilitato': utente_abilitato })
    else:
        utente_abilitato = False
    return render(request, 'inipec_massiva.html', { 'utente_abilitato': utente_abilitato })


def registro_imprese_get_bearer():
    parametri_registro_imprese = RegistroImpreseParametri.objects.get(id=1)
    kid = parametri_registro_imprese.kid
    alg = parametri_registro_imprese.alg
    typ = parametri_registro_imprese.typ
    issuer = parametri_registro_imprese.iss
    subject = parametri_registro_imprese.sub
    aud = parametri_registro_imprese.aud
    purposeId = parametri_registro_imprese.purposeid
    audience = parametri_registro_imprese.audience
    baseurlauth = parametri_registro_imprese.baseurlauth
    target = parametri_registro_imprese.target
    clientid = parametri_registro_imprese.clientid


    private_key = parametri_registro_imprese.private_key

    issued = datetime.datetime.utcnow()
    delta = datetime.timedelta(minutes=43200)
    expire_in = issued + delta
    jti = uuid.uuid4()
    headers_rsa = {
        "kid": kid,
        "alg": alg,
        "typ": typ
    }

    payload = {
        "iss": issuer,
        "sub": subject,
        "aud": aud,
        "purposeId": purposeId,
        "jti": str(jti),
        "iat": issued,
        "exp": expire_in
    }

    asserzione = jwt.encode(payload, private_key, algorithm="RS256", headers=headers_rsa)

    curl_command = (
        f"curl --location --silent --request POST {baseurlauth}/token.oauth2 "
        f"--header 'Content-Type: application/x-www-form-urlencoded' "
        f"--data-urlencode 'client_id={issuer}' "
        f"--data-urlencode 'client_assertion={asserzione}' "
        "--data-urlencode 'client_assertion_type=urn:ietf:params:oauth:client-assertion-type:jwt-bearer' "
        "--data-urlencode 'grant_type=client_credentials'"
    )

    if not sys.platform.startswith('linux'):
        curl_command=curl_command.replace("'", '"')
    result = subprocess.run(curl_command, shell=True, capture_output=True, text=True)
    data = json.loads(result.stdout)
    return data['access_token']


def registro_imprese(request):
    if request.user.id:
        utente_sessione = UtentiParametri.objects.get(id=request.user.id)
        utente_abilitato = utente_sessione.registro_imprese
        if request.method == 'POST':
            testi_registro = [
                ('c_fonte', 'Codice fonte'),
                ('fonte', 'Fonte'),
                ('descrizione_tipo_impresa', 'Tipo Impresa'),
                ('tipo_soggetto', 'Tipo soggetto'),
                ('descrizione_tipo_soggetto', 'Descrizione tipo soggetto'),
                ('tipo_impresa', 'Tipo impresa'),
                ('c_causale_cess', 'Codice causale cessazione/cancellazione'),
                ('dt_iscrizione_ri', 'Data di iscrizione al Registro Imprese'),
                ('dt_iscrizione_rea', 'Data di iscrizione al REA'),
                ('dt_atto_costituzione', 'Data atto di costituzione della società'),
                ('dt_costituzione', 'Data di costituzione dell\'impresa individuale'),
                ('dt_cancellazione', 'Data di cancellazione dal Registro Imprese'),
                ('dt_cessazione', 'Data di cessazione'),
                ('causale_cess', 'Codice causale cessazione/cancellazione'),
                ('denominazione', 'Denominazione dell\'impresa'),
                ('c_fiscale', 'Codice fiscale dell\'impresa'),
                ('partita_iva', 'Partita IVA'),
                ('cciaa', 'Camera di Commercio di iscrizione dell\'impresa'),
                ('n_rea', 'Numero REA dell\'impresa se iscritta al R.I.'),
                ('n_rd', 'Numero Registro Ditte se la ditta è del Registro Ditte'),
                ('stato_impresa', 'Stato impresa'),
                ('indirizzo_posta_certificata', 'Domicilio digitale'),
                ('stato_ditta', 'Stato impresa'),
                ('c_codifica', 'Codifica'),
                ('codifica', 'Tipo Classificazione'),
                ('classificazioni_ateco', 'Classificazioni_ATECO'),
                ('c_importanza', 'Codice_importanza'),
                ('importanza', 'Importanza'),
                ('forma_giuridica', 'Forma_giuridica'),
                ('indirizzo_localizzazione', 'Indirizzo'),
                ('dt_inizio_attivita_impresa', 'Data inizio attività'),
                ('dt_inizio', 'Data inizio attività'),
                ('attivita', 'Attività'),
                ('c_attivita', 'Codice attività'),
                ('c_stato', 'Indicatore stato dell\'attività: I (Inattiva), S (Sospesa)'),
                ('stato', 'Descrizione attributo precedente'),
                ('attivita_prevalente', 'Descrizione dell\'attività prevalente'),
                ('attivita_esercitata', 'Descrizione dell\'attività esercitata'),
                ('attivita_secondaria_esercitata', 'Descrizione attività secondaria esercitata'),
                ('f_presenza_pco_s', 'Presenza di procedure concorsuali in corso o pregresse'),
                ('dt_soci_titolari_al', 'Data di estrazione della visura CAD'),
                ('f_consorzio', 'Elenco di consorziati'),
                ('capitale_sociale', 'Capitale sociale'),
                ('c_valuta', 'Codice valuta'),
                ('valuta', 'Valuta'),
                ('deliberato', 'Deliberato'),
                ('ammontare', 'Ammontare'),
                ('sottoscritto', 'Sottoscritto'),
                ('versato', 'Versato'),
                ('sistema_amministrazione', 'Sistema ammistrazione'),
                ('forme_amministrative', 'Forma di amministrazione'),
                ('soggetto_controllo_contabile', 'Soggetto con controllo stabile'),
                ('collegio_sindacale', 'Collegio sindacale'),
                ('dt_termine', 'Data termine'),
                ('scadenza_esercizi', 'Scadenza esercizi'),
                ('dt_primo_esercizio', 'Data primo esercizio'),
                ('esercizi_successivi', 'Esercizi successivi'),
                ('giorni_proroga_bilancio', 'Giorni di proroga bilancio'),
                ('riquadro', 'Riquadro'),
                ('composizione_quote', 'Composizione quote'),
                ('titolari', 'Titolari'),
                ('titolare', 'Titolare'),
                ('anagrafica_titolare', 'Anagrafica titolare'),
                ('cognome', 'Cognome'),
                ('nome', 'Nome'),
                ('diritto_partecipazione', 'Diritto partecipazione'),
                ('tipo', 'Tipo'),
                ('c_tipo', 'Codice tipo'),
                ('Codice_importanza', 'Codice importanza'),
                ('ruolo_ridotto', 'Ruolo ridotto'),
                ('c_ente_rilascio','Codice ente rilascio'),
                ('ente_rilascio','Ente rilascui'),
                ('data','Data'),
                ('atti_conferimento_cariche', ''),
                ('atto_conferimento_cariche', ''),
                ('cariche', 'Cariche'),
                ('carica', 'Carica'),
                ('c_carica', 'Codice carica'),
                ('poteri_persona', 'Poteri persona'),
                ('p_poteri', 'Codice potere'),
                ('text', 'Testo'),
                ('persona_fisica', 'Persona fisica'),
                ('progressivo', 'Progressivo'),
                ('dt','Data'),
                ('estremi_nascita','Estremi di nascita'),
                ('frazione_numeratore','Frazione numeratore'),
                ('frazione_denominatore','Frazione denominatore'),
                ('valore_nominale', 'Valore nominale'),
                ('f_rappresentante_ri','f_rappresentante_ri'),
                ('dt_apertura','Data apertura'),
                ('comune','Comune'),
                ('provincia','Provincia'),
                ('toponimo','Toponimo'),
                ('cap','CAP'),
                ('Classificazioni_ATECO','Classificazioni ATECO'),
                ('Codice_importanza','Codice importanza'),
                ]
            data = []
            dati_identificativi = ""
            info_attivita = ""
            albi_ruoli_licenze_ridotti = ""
            persone_sede = ""
            localizzazioni = ""
            elenco_soci = ""
            info_statuto = ""
            amministrazione_controllo = ""
            info_patrimoniali_finanziarie = ""
            scritta_pco_s = ""

            cf = request.POST.get('input_CF')
            data.append(cf)
            bearer = registro_imprese_get_bearer()
            correttezza_cf_azienda = verifica_cf_azienda(cf)
            if correttezza_cf_azienda == 1:
                elenco_dati_registro_temp = registro_imprese_verifica_utente(cf, bearer)
                if elenco_dati_registro_temp :
                    elenco_dati_registro = sostituisci_chiavi(elenco_dati_registro_temp, testi_registro)
                    elenco_dati_registro = converti_data(elenco_dati_registro)
                    data.append(elenco_dati_registro)
                    if 'dati_identificativi' in elenco_dati_registro['blocchi_impresa'] :
                        dati_identificativi = elenco_dati_registro['blocchi_impresa']['dati_identificativi']
                    if 'info_attivita' in elenco_dati_registro['blocchi_impresa'] :
                        info_attivita = elenco_dati_registro['blocchi_impresa']['info_attivita']
                    if 'albi_ruoli_licenze_ridotti' in elenco_dati_registro['blocchi_impresa'] :
                        albi_ruoli_licenze_ridotti = elenco_dati_registro['blocchi_impresa']['albi_ruoli_licenze_ridotti']
                    if 'persone_sede' in elenco_dati_registro['blocchi_impresa'] :
                        persone_sede = elenco_dati_registro['blocchi_impresa']['persone_sede']
                    if 'localizzazioni' in elenco_dati_registro['blocchi_impresa'] :
                        localizzazioni = elenco_dati_registro['blocchi_impresa']['localizzazioni']
                    if 'elenco_soci' in elenco_dati_registro['blocchi_impresa'] :
                        elenco_soci = elenco_dati_registro['blocchi_impresa']['elenco_soci']
                    if 'info_statuto' in elenco_dati_registro['blocchi_impresa'] :
                        info_statuto = elenco_dati_registro['blocchi_impresa']['info_statuto']
                    if 'amministrazione_controllo' in elenco_dati_registro['blocchi_impresa'] :
                        amministrazione_controllo = elenco_dati_registro['blocchi_impresa']['amministrazione_controllo']
                    if 'info_patrimoniali_finanziarie' in elenco_dati_registro['blocchi_impresa'] :
                        info_patrimoniali_finanziarie = elenco_dati_registro['blocchi_impresa']['info_patrimoniali_finanziarie']
                    if 'scritta_pco_s' in elenco_dati_registro['blocchi_impresa'] :
                        scritta_pco_s = elenco_dati_registro['blocchi_impresa']['scritta_pco_s']
                else:
                    data.append("Ditta non presente")
            else:
                data.append("Codice fiscale non corretto")

            salva_log(request.user,"Verifica Registro Imprese", "Verificato azienda " + cf)
            return render(request, 'registro_imprese.html', {'data': data, 'dati_identificativi' : dati_identificativi, 'info_attivita' : info_attivita, 'albi_ruoli_licenze_ridotti' : albi_ruoli_licenze_ridotti, 'persone_sede' : persone_sede, 'localizzazioni' : localizzazioni, 'elenco_soci' : elenco_soci, 'info_statuto' : info_statuto, 'amministrazione_controllo' : amministrazione_controllo, 'info_patrimoniali_finanziarie' : info_patrimoniali_finanziarie, 'scritta_pco_s' : scritta_pco_s, 'utente_abilitato': utente_abilitato })

    else:
        utente_abilitato = False

    return render(request, 'registro_imprese.html', { 'utente_abilitato': utente_abilitato })


def registro_imprese_verifica_utente(cf, bearer):
    registro_imprese_parametri = RegistroImpreseParametri.objects.get(id=1)
    url = registro_imprese_parametri.target + '/dettaglio/codicefiscale' ####da vedere
    #### /ricerca/codicefiscale
    #### /dettaglio/codicefiscale

    headers = {
        "Authorization": f"Bearer {bearer}"
        }
    params = {
        "codiceFiscale": cf
        }
    response = requests.get(url, headers=headers, params=params)

    def modify_keys(d):
        new_d = {}
        for key, value in d.items():
            if isinstance(value, dict):
                value = modify_keys(value)
            elif isinstance(value, list):
                value = [modify_keys(item) if isinstance(item, dict) else item for item in value]
            new_key = key.replace("-", "_").replace("@", "").replace("#", "")
            new_d[new_key] = value
        return new_d

    if response.status_code == 200:
        xml_data = response.content
        parsed_data = xmltodict.parse(xml_data)
        modified_data = modify_keys(parsed_data)
        return modified_data
    else:
        return False


def registro_imprese_valida_cf(cf):
    numeri = [int(c) for c in cf]
    somma = 0
    # Posizioni dispari: 1,3,5,7,9 → indici 0,2,4,6,8
    for i in [0, 2, 4, 6, 8]:
        somma += numeri[i]
    # Posizioni pari: 2,4,6,8,10 → indici 1,3,5,7,9
    for i in [1, 3, 5, 7, 9]:
        val = numeri[i] * 2
        if val > 9:
            val -= 9
        somma += val
    # Calcolo cifra di controllo
    controllo = (10 - (somma % 10)) % 10
    return controllo == numeri[10]


def sostituisci_chiavi(json_data, mapping_array):
    if isinstance(json_data, dict):
        new_dict = {}
        for chiave, valore in json_data.items():
            if isinstance(valore, dict):
                valore = sostituisci_chiavi(valore, mapping_array)
            elif isinstance(valore, list):
                valore = [sostituisci_chiavi(item, mapping_array) for item in valore]
            for item in mapping_array:
                if chiave == item[0]:
                    chiave = item[1]
                    break
            new_dict[chiave] = valore
        return new_dict
    elif isinstance(json_data, list):
        return [sostituisci_chiavi(item, mapping_array) for item in json_data]
    else:
        return json_data


def impostazioni_registro_imprese(request):
    parametri_registro_imprese = RegistroImpreseParametri.objects.all()
    if request.method == 'POST':
        kid = request.POST.get('kid')
        alg = request.POST.get('alg')
        typ = request.POST.get('typ')
        iss = request.POST.get('iss')
        sub = request.POST.get('sub')
        aud = request.POST.get('aud')
        purposeid = request.POST.get('purposeid')
        audience = request.POST.get('audience')
        baseurlauth = request.POST.get('baseurlauth')
        target = request.POST.get('target')
        clientid = request.POST.get('clientid')
        private_key = request.POST.get('private_key')
        ver_eservice = request.POST.get('ver_eservice')

        dati = RegistroImpreseParametri(1, kid, alg, typ, iss, sub, aud, purposeid, audience, baseurlauth, target, clientid, private_key, ver_eservice)
        dati.save()
        salva_log(request.user,"Impostazioni Registro Imprese", "modifica parametri")
    else:
        parametri_registro_imprese = RegistroImpreseParametri.objects.all()

    return render(request, 'impostazioni_registro_imprese.html', {'parametri_registro_imprese': parametri_registro_imprese})

