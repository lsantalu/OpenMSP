from django.shortcuts import render
from datetime import date

from .utils import salva_log
from .utils import converti_data


import io
import csv
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment
from openpyxl.utils import get_column_letter
from django.http import HttpResponse


def verifica_cf(request):
    if len(request) != 16:
        return -1  # Lunghezza errata

    set1 = "0123456789ABCDEFGHIJKLMNOPQRSTUVWXYZ"
    set2 = "ABCDEFGHIJABCDEFGHIJKLMNOPQRSTUVWXYZ"
    setpari = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"
    setdisp = "BAKPLCQDREVOSFTGUHMINJWZYX"
    mesi = {'A': 1, 'B': 2, 'C': 3, 'D': 4, 'E': 5, 'H': 6,'L': 7, 'M': 8, 'P': 9, 'R': 10, 'S': 11, 'T': 12 }
    s = 0
    cf=request.upper()
    for i in range(1, 14, 2):
        if cf[i].isupper() or cf[i].islower() or cf[i].isdigit():
            s += setpari.index(set2[set1.index(cf[i])])
        else:
            return -1 #CODICE FISCALE SBAGLIATO
    for i in range(0, 15, 2):
        if cf[i].isupper() or cf[i].islower() or cf[i].isdigit():
            s += setdisp.index(set2[set1.index(cf[i])])
        else:
            return -1 #CODICE FISCALE SBAGLIATO

    try:
        if mesi[cf[8:9]]:
            month = mesi[cf[8:9]]
        pass
    except KeyError:
        return -1 #CODICE FISCALE SBAGLIATO

    if not s % 26 != ord(cf[15]) - ord('A'): #CODICE FISCALE CORRETTO
        year = int(cf[6:8])
        if year > date.today().year % 100:
            year += 1900
        else:
            year += 2000
        day = int(cf[9:11])
        if day > 40:
            day = day - 40
        birth_date = date(year, month, day)
        age = date.today() - birth_date
        
        if age.days < 365 * 18:
            return 2 #CODICE FISCALE MINORENNE
        return 1 #CODICE FISCALE MAGGIORENNE
    else:
        return -1 #CODICE FISCALE SBAGLIATO


def verifica_cf_azienda(request):
    ##CF persone e P_IVA
    if len(request) == 16:
        return verifica_cf(request)
    else:
        if len(request) != 11:
            return -1  # Lunghezza errata
    try:
        partita_iva_int = int(request)
    except ValueError:
        return -1  # Contiene caratteri non numerici

    if partita_iva_int == 0:
        return -1  # Partita IVA non può essere 0

    s = 0
    for i, c in enumerate(request):
        if i % 2 == 0:
            s += int(c)
        else:
            double = int(c) * 2
            s += double if double < 10 else double - 9

    return 1 if s % 10 == 0 else -1  # Verifica di correttezza        


def verifica_cf_export_excel(request):
    data = request.session.get("multi_data", [])  # Oppure recuperalo come preferisci
    wb = Workbook()
    ws = wb.active
    ws.append(["#", "CF", "Verifica"])

    # Dizionario per la larghezza massima di ogni colonna
    max_lengths = [len(cell.value) for cell in ws[1]]  # Larghezze iniziali dall'intestazione
    
    if not data:
        return HttpResponse("Nessun dato disponibile per l'esportazione", status=400)
    
    for i, row in enumerate(data, 1):
        split_row = row.split()
        stato = split_row[1]
        color = "FFFFFF"  # default bianco

        if stato == "1":
            color = "C6EFCE"  # verde chiaro
        elif stato == "2":
            color = "FFEB9C"  # giallo
        else:
            color = "FFC7CE"  # rosso

        fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

        ws.append([i, split_row[0], split_row[1]])
        for col in ws.iter_rows(min_row=ws.max_row, max_row=ws.max_row):
            for i, cell in enumerate(col):
                cell.fill = fill
                value_length = len(str(cell.value)) if cell.value else 0
                if len(max_lengths) <= i:
                    max_lengths.append(value_length)
                else:
                    max_lengths[i] = max(max_lengths[i], value_length)

    # Imposta larghezza colonne
    for i, width in enumerate(max_lengths, 1):
        col_letter = get_column_letter(i)
        ws.column_dimensions[col_letter].width = width

    # Scrive l'Excel su un buffer
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=EsitoVerificaCF.xlsx'
    return response



def verifica_cf_massivo(request):
    if request.method == 'POST':
        csv_file = request.FILES['cf_csv']
        data = []
        contatore = 0
        if csv_file.name.endswith('.csv'):
            csv_file_text = io.TextIOWrapper(csv_file.file, encoding='utf-8')
            csv_reader = csv.reader(csv_file_text)
            for row in csv_reader:
                if row[0]:  # Se row[0] non è vuoto o None
                    data.append(row[0].strip().upper() +" "+ str(verifica_cf(row[0].strip().upper())))
                    contatore += 1
            if request.user.id:
                salva_log(request.user,"Verifica correttezza CF massivo", "Verificati n. " + str(contatore) + " CF")
            request.session["multi_data"] = data  # <--- Salva i dati nella sessione
            return render(request, 'verifica_cf.html', {'data': data})
        elif csv_file.name.endswith('.xlsx') or csv_file.name.endswith('.XLSX') or csv_file.name.endswith('.Xlsx'):
            wb = openpyxl.load_workbook(csv_file)
            sheet = wb.active
            for row in sheet.iter_rows(min_row=1, values_only=True):
                if row[0]:
                    data.append(row[0].strip().upper() +" "+ str(verifica_cf(row[0].strip().upper())))
                    contatore += 1
            if request.user.id:
                salva_log(request.user,"Verifica correttezza CF massivo", "Verificati n. " + str(contatore) + " CF")
            request.session["multi_data"] = data  # <--- Salva i dati nella sessione
            return render(request, 'verifica_cf.html', {'data': data})         
        else:
            if request.user.id:
                salva_log(request.user,"Verifica correttezza CF massivo", "Errore caricamento file CSV")
            return render(request, 'verifica_cf.html', {'error': 'Il file non è un CSV'})
    return render(request, 'verifica_cf.html')


def verifica_cf_aziende_massivo(request):
    if request.method == 'POST':
        csv_file = request.FILES['cf_csv']
        data = []
        contatore = 0
        if csv_file.name.endswith('.csv'):
            csv_file_text = io.TextIOWrapper(csv_file.file, encoding='utf-8')
            csv_reader = csv.reader(csv_file_text)
            for row in csv_reader:
                if row[0]:  # Se row[0] non è vuoto o None
                    if len(row[0]) in [8, 9, 10]:
                        row[0] = row[0].zfill(11)
                    data.append(row[0].strip().upper() +" "+ str(verifica_cf_azienda(row[0].strip().upper())))
                    contatore += 1
            if request.user.id:
                salva_log(request.user,"Verifica correttezza CF aziende massivo", "Verificati n. " + str(contatore) + " CF")
            request.session["multi_data"] = data  # <--- Salva i dati nella sessione
            return render(request, 'verifica_cf_aziende.html', {'data': data})
        elif csv_file.name.endswith('.xlsx') or csv_file.name.endswith('.XLSX') or csv_file.name.endswith('.Xlsx'):
            wb = openpyxl.load_workbook(csv_file)
            sheet = wb.active
            for row in sheet.iter_rows(min_row=1, values_only=True):
                if row[0]:
                    row = list(row)
                    if not isinstance(row[0], str):
                        row[0] = str(row[0])
                    if len(row[0]) in [8, 9, 10]:
                        row[0] = row[0].zfill(11)
                    data.append(row[0].strip().upper() +" "+ str(verifica_cf_azienda(row[0].strip().upper())))
                    contatore += 1
                if request.user.id:
                    salva_log(request.user,"Verifica correttezza CF aziende massivo", "Verificati n. " + str(contatore) + " CF")
            request.session["multi_data"] = data  # <--- Salva i dati nella sessione
            return render(request, 'verifica_cf_aziende.html', {'data': data})                  
        else:
            if request.user.id:
                salva_log(request.user,"Verifica correttezza CF aziende massivo", "Errore caricamento file CSV")
            return render(request, 'verifica_cf_aziende.html', {'error': 'Il file non è un CSV'})
    return render(request, 'verifica_cf_aziende.html')


def verifica_cf_aziende_export_excel(request):
    data = request.session.get("multi_data", [])  # Oppure recuperalo come preferisci
    wb = Workbook()
    ws = wb.active
    ws.append(["#", "CF", "Verifica"])

    # Dizionario per la larghezza massima di ogni colonna
    max_lengths = [len(cell.value) for cell in ws[1]]  # Larghezze iniziali dall'intestazione
    
    if not data:
        return HttpResponse("Nessun dato disponibile per l'esportazione", status=400)
    
    for i, row in enumerate(data, 1):
        split_row = row.split()
        stato = split_row[1]
        color = "FFFFFF"  # default bianco
        if stato == "1":
            color = "C6EFCE"  # verde chiaro
        else:
            color = "FFC7CE"  # rosso

        fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

        ws.append([i, split_row[0], split_row[1]])
        for col in ws.iter_rows(min_row=ws.max_row, max_row=ws.max_row):
            for i, cell in enumerate(col):
                cell.fill = fill
                value_length = len(str(cell.value)) if cell.value else 0
                if len(max_lengths) <= i:
                    max_lengths.append(value_length)
                else:
                    max_lengths[i] = max(max_lengths[i], value_length)

    # Imposta larghezza colonne
    for i, width in enumerate(max_lengths, 1):
        col_letter = get_column_letter(i)
        ws.column_dimensions[col_letter].width = width

    # Scrive l'Excel su un buffer
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=EsitoVerificaCFAziende.xlsx'
    return response

