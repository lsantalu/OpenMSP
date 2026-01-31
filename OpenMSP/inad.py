from django.shortcuts import render
from impostazioni.models import InadParametri
from impostazioni.models import UtentiParametri

from .utils import salva_log
from .verifica_cf import verifica_cf

import datetime
import uuid
import jwt
import subprocess
import json
import io
import csv
import sys
import re
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment
from openpyxl.utils import get_column_letter

from django.http import HttpResponse


def inad_export_excel(request):
    data = request.session.get("multi_data", [])  # Oppure recuperalo come preferisci
    wb = Workbook()
    ws = wb.active
    ws.append(["#", "CF", "Verifica", "Domicilio digitale"])

    # Dizionario per la larghezza massima di ogni colonna
    max_lengths = [len(cell.value) for cell in ws[1]]  # Larghezze iniziali dall'intestazione
    
    if not data:
        return HttpResponse("Nessun dato disponibile per l'esportazione", status=400)
    
    for i, row in enumerate(data, 1):
        split_row = row.split()
        stato = split_row[1]
        color = "FFFFFF"  # default bianco

        if stato == "1" and "@" in split_row[2]:
            color = "C6EFCE"  # verde chiaro
        elif stato == "2":
            color = "FFEB9C"  # giallo
        else:
            color = "FFC7CE"  # rosso

        fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

        ws.append([i, split_row[0], split_row[1], " ".join(split_row[2:])])
        for col in ws.iter_rows(min_row=ws.max_row, max_row=ws.max_row):
            for i, cell in enumerate(col):
                cell.fill = fill
                value_length = len(str(cell.value)) if cell.value else 0
                if len(max_lengths) <= i:
                    max_lengths.append(value_length)
                else:
                    max_lengths[i] = max(max_lengths[i], value_length)

    # Imposta larghezza colonne
    for i, width in enumerate(max_lengths, 1):
        col_letter = get_column_letter(i)
        ws.column_dimensions[col_letter].width = width

    # Scrive l'Excel su un buffer
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=EsitoDomiciliDigitaliPF.xlsx'
    return response


def inad_get_bearer():
    parametri_inad = InadParametri.objects.get(id=1)
    kid = parametri_inad.kid
    alg = parametri_inad.alg
    typ = parametri_inad.typ
    issuer = parametri_inad.iss
    subject = parametri_inad.sub
    aud = parametri_inad.aud
    purposeId = parametri_inad.purposeid
    audience = parametri_inad.audience
    baseurlauth = parametri_inad.baseurlauth
    target = parametri_inad.target
    clientid = parametri_inad.clientid
    private_key = parametri_inad.private_key
    
    
    issued = datetime.datetime.utcnow()
    delta = datetime.timedelta(minutes=43200)
    expire_in = issued + delta
    jti = uuid.uuid4()
    headers_rsa = {
        "kid": kid,
        "alg": alg,
        "typ": typ
    }

    payload = {
        "iss": issuer,
        "sub": subject,
        "aud": aud,
        "purposeId": purposeId,
        "jti": str(jti),
        "iat": issued,
        "exp": expire_in
    }
    
    asserzione = jwt.encode(payload, private_key, algorithm="RS256", headers=headers_rsa)

    curl_command = (
        f"curl --location --silent --request POST {baseurlauth}/token.oauth2 "
        f"--header 'Content-Type: application/x-www-form-urlencoded' "
        f"--data-urlencode 'client_id={issuer}' "
        f"--data-urlencode 'client_assertion={asserzione}' "
        "--data-urlencode 'client_assertion_type=urn:ietf:params:oauth:client-assertion-type:jwt-bearer' "
        "--data-urlencode 'grant_type=client_credentials'"
    )

    if not sys.platform.startswith('linux'):
        curl_command=curl_command.replace("'", '"')

    result = subprocess.run(curl_command, shell=True, capture_output=True, text=True)
    data = json.loads(result.stdout)
    return data['access_token']


def inad_singola(request):
    if request.user.id:
        utente_sessione = UtentiParametri.objects.get(id=request.user.id)
        utente_abilitato = utente_sessione.inad_singolo
        if request.method == 'POST':
            data = []
            cf = request.POST.get('input_CF')
            data.append(cf)
            bearer = inad_get_bearer()
            correttezza_cf = verifica_cf(cf) 
            if correttezza_cf == 1:
                parsed_output= inad_verifica_utente(cf, bearer)
                data.append(estrai_mail(json.dumps(parsed_output)))
            elif correttezza_cf == 2:
                data.append("Codice fiscale di persona minorenne") 
            else:
                data.append("Codice fiscale non corretto")
            salva_log(request.user,"Verifica INAD singolo", "Verificato domicilio utente " + cf )
            return render(request, 'inad_singola.html', {'data': data, 'utente_abilitato': utente_abilitato })
    else: 
        utente_abilitato = False
    return render(request, 'inad_singola.html', { 'utente_abilitato': utente_abilitato })


def inad_massiva(request):
    if request.user.id:
        utente_sessione = UtentiParametri.objects.get(id=request.user.id)
        utente_abilitato = utente_sessione.inad_massivo
        if request.method == 'POST':
            csv_file = request.FILES['cf_csv']
            data = []
            contatore = 0
            bearer = inad_get_bearer()
            if csv_file.name.endswith('.csv'):
                csv_file_text = io.TextIOWrapper(csv_file.file, encoding='utf-8')
                csv_reader = csv.reader(csv_file_text)
                for row in csv_reader:
                    if row[0]:  # Se row[0] non è vuoto o None
                        correttezza_cf = verifica_cf(row[0].strip().upper())
                        if correttezza_cf == 1:
                            parsed_output= inad_verifica_utente(row[0].strip().upper(), bearer)
                            data.append(row[0].strip().upper() + " " + str(correttezza_cf) + " " + estrai_mail(json.dumps(parsed_output)))
                        elif correttezza_cf == 2:
                            data.append(row[0].strip().upper() + " " + str(correttezza_cf) + " Codice fiscale di persona minorenne")
                        else:
                            data.append(row[0].strip().upper() + " " + str(correttezza_cf) + " Codice fiscale non corretto")               
                        contatore += 1
                salva_log(request.user,"Verifica INAD massivo", "Verificati n. " + str(contatore) + " CF")
                request.session["multi_data"] = data  # <--- Salva i dati nella sessione
                return render(request, 'inad_massiva.html', {'data': data, 'utente_abilitato': utente_abilitato })
            elif csv_file.name.endswith('.xlsx') or csv_file.name.endswith('.XLSX') or csv_file.name.endswith('.Xlsx'):
                wb = openpyxl.load_workbook(csv_file)
                sheet = wb.active
                for row in sheet.iter_rows(min_row=1, values_only=True):
                    if row[0]:
                        correttezza_cf = verifica_cf(row[0].strip().upper())
                        if correttezza_cf == 1:
                            parsed_output= inad_verifica_utente(row[0].strip().upper(), bearer)
                            data.append(row[0].strip().upper() + " " + str(correttezza_cf) + " " + estrai_mail(json.dumps(parsed_output)))
                        elif correttezza_cf == 2:
                            data.append(row[0].strip().upper() + " " + str(correttezza_cf) + " Codice fiscale di persona minorenne")
                        else:
                            data.append(row[0].strip().upper() + " " + str(correttezza_cf) + " Codice fiscale non corretto")               
                        contatore += 1
                salva_log(request.user,"Verifica INAD massivo", "Verificati n. " + str(contatore) + " CF")
                request.session["multi_data"] = data  # <--- Salva i dati nella sessione
                return render(request, 'inad_massiva.html', {'data': data, 'utente_abilitato': utente_abilitato })           
            else:
                salva_log(request.user,"Verifica INAD massivo", "Errore caricamento file CSV")
                return render(request, 'inad_massiva.html', {'error': 'Il file non è un CSV', 'utente_abilitato': utente_abilitato })
    else: 
        utente_abilitato = False
    return render(request, 'inad_massiva.html', { 'utente_abilitato': utente_abilitato })


def inad_verifica_utente(cf, bearer):
    inad_parametri = InadParametri.objects.get(id=1)
    url = inad_parametri.target + '/extract'
    curl_command = (
        f'curl --silent --request GET '
        f"--url '{url}/{cf}?practicalReference=ABC123' "
        f"--header 'Authorization: Bearer {bearer}'"
        )
    if not sys.platform.startswith('linux'):
        curl_command=curl_command.replace("'", '"')    
            
    result = subprocess.run(curl_command, shell=True, capture_output=True, text=True) 
    return json.loads(result.stdout)


def impostazioni_inad(request):
    parametri_inad = InadParametri.objects.all()
    if request.method == 'POST':
        kid = request.POST.get('kid')
        alg = request.POST.get('alg')
        typ = request.POST.get('typ')
        iss = request.POST.get('iss')
        sub = request.POST.get('sub')
        aud = request.POST.get('aud')
        purposeid = request.POST.get('purposeid')
        audience = request.POST.get('audience')
        baseurlauth = request.POST.get('baseurlauth')
        target = request.POST.get('target')
        clientid = request.POST.get('clientid')
        private_key = request.POST.get('private_key')
        ver_eservice = request.POST.get('ver_eservice')
        
        
        
        dati = InadParametri(1, kid, alg, typ, iss, sub, aud, purposeid, audience, baseurlauth, target, clientid, private_key, ver_eservice)
        dati.save()
        salva_log(request.user,"Impostazioni INAD", "modifica parametri")
    else:
        parametri_inad = InadParametri.objects.all()
            
    return render(request, 'impostazioni_inad.html', {'parametri_inad': parametri_inad})

def estrai_mail(indirizzo):
    if "NOT_FOUND" in indirizzo:
        return "Domicilio non presente"
    else:
        regex_email = r'[A-Z0-9._%+-]+@[A-Z0-9.-]+\.[A-Z]{2,4}'
        indirizzi = re.findall(regex_email, indirizzo, re.IGNORECASE)
        return indirizzi[0]