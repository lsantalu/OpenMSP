from django.shortcuts import render

from impostazioni.models import UtentiParametri
from impostazioni.models import ServiziParametri
from impostazioni.models import AnisServizi
from impostazioni.models import AnisParametri

from .utils import salva_log
from .utils import converti_data
from .verifica_cf import verifica_cf

##from datetime import datetime, date
import datetime
from jose.constants import Algorithms
import http.client, urllib.parse
import hashlib
import random
import base64
import datetime
import uuid
import jwt
###import subprocess
import requests
import json
import io
import csv
import re
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

from django.http import HttpResponse

def anis_iscrizioni_export_excel(request):
    data = request.session.get("multi_data", [])  # Oppure recuperalo come preferisci
    wb = Workbook()
    ws = wb.active
    ws.append(["Codice Fiscale", "Istituto", "Tipologia corso", "Nome corso", "Classe", "Anno accademico", "Durata corso"])

    # Dizionario per la larghezza massima di ogni colonna
    max_lengths = [len(cell.value) for cell in ws[1]]  # Larghezze iniziali dall'intestazione

    if not data:
        return HttpResponse("Nessun dato disponibile per l'esportazione", status=400)

    for i, row in enumerate(data, 1):
        color = "FFFFFF"  # default bianco

        if isinstance(row, dict):
            cf = row['personal_data']['tax_code']
            if len(row.get("enrollments")) == 0 :
                color = "FFC7CE"
                ws.append([cf, "La richiesta effettuata non produce alcun risultato", "N/A", "N/A", "N/A", "N/A", "N/A"])
            else:
                for enrol in row['enrollments']:
                    color = "C6EFCE"  # verde chiaro
                    istituto = enrol['institute_name'] + " (" + enrol['institute_code'] + ")"
                    tipo_corso = enrol['programme_type_code']
                    nome_corso = enrol['degree_course_code']
                    classe = enrol['degree_class_code']
                    anno_accademico = enrol['academic_year']
                    durata_corso = enrol['degree_course_year']
                    ws.append([cf, istituto, tipo_corso, nome_corso, classe, anno_accademico, durata_corso])
        else:
            split_row = row.split()
            cf = split_row[0]
            if "Codice" in split_row[1]:
                color = "FFC7CE"  # rosso
                ws.append([cf, "Codice fiscale non corretto", "N/A", "N/A", "N/A", "N/A", "N/A"])

        fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        for col in ws.iter_rows(min_row=ws.max_row, max_row=ws.max_row):
            for i, cell in enumerate(col):
                cell.fill = fill
                value_length = len(str(cell.value)) if cell.value else 0
                if len(max_lengths) <= i:
                    max_lengths.append(value_length)
                else:
                    max_lengths[i] = max(max_lengths[i], value_length)

    # Imposta larghezza colonne
    for i, width in enumerate(max_lengths, 1):
        col_letter = get_column_letter(i)
        ws.column_dimensions[col_letter].width = width

    # Scrive l'Excel su un buffer
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=EsitoAnisIscrizioni.xlsx'
    return response


def anis_iscrizioni_export_csv(request):
    data = request.session.get("multi_data", [])
    
    response = HttpResponse(
        content_type='text/csv',
        headers={'Content-Disposition': 'attachment; filename="EsitoAnisIscrizioni.csv"'},
    )

    writer = csv.writer(response, delimiter=';')
    writer.writerow(["Codice Fiscale", "Istituto", "Tipologia corso", "Nome corso", "Classe", "Anno accademico", "Durata corso"])

    if not data:
        return HttpResponse("Nessun dato disponibile per l'esportazione", status=400)

    for row in data:
        if isinstance(row, dict):
            cf = row['personal_data']['tax_code']
            if len(row.get("enrollments")) == 0:
                writer.writerow([cf, "La richiesta effettuata non produce alcun risultato", "N/A", "N/A", "N/A", "N/A", "N/A"])
            else:
                for enrol in row['enrollments']:
                    istituto = enrol['institute_name'] + " (" + enrol['institute_code'] + ")"
                    tipo_corso = enrol['programme_type_code']
                    nome_corso = enrol['degree_course_code']
                    classe = enrol['degree_class_code']
                    anno_accademico = enrol['academic_year']
                    durata_corso = enrol['degree_course_year']
                    writer.writerow([cf, istituto, tipo_corso, nome_corso, classe, anno_accademico, durata_corso])
        else:
            split_row = row.split()
            cf = split_row[0]
            if "Codice" in split_row[1]:
                writer.writerow([cf, "Codice fiscale non corretto", "N/A", "N/A", "N/A", "N/A", "N/A"])

    return response


def anis_titoli_export_excel(request):
    data = request.session.get("multi_data", [])  # Oppure recuperalo come preferisci
    wb = Workbook()
    ws = wb.active
    ws.append(["Codice Fiscale", "Istituto", "Qualifica", "Tipologia corso", "Nome corso", "Classe", "Data conseguimento", "Valutazione"])

    # Dizionario per la larghezza massima di ogni colonna
    max_lengths = [len(cell.value) for cell in ws[1]]  # Larghezze iniziali dall'intestazione

    if not data:
        return HttpResponse("Nessun dato disponibile per l'esportazione", status=400)

    for i, row in enumerate(data, 1):
        color = "FFFFFF"  # default bianco

        if isinstance(row, dict):
            cf = row['personal_data']['tax_code']
            if len(row.get("qualifications")) == 0 :
                color = "FFC7CE"
                ws.append([cf, "La richiesta effettuata non produce alcun risultato", "N/A", "N/A", "N/A", "N/A", "N/A", "N/A"])
                fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                for col in ws.iter_rows(min_row=ws.max_row, max_row=ws.max_row):
                    for i, cell in enumerate(col):
                        cell.fill = fill
                        value_length = len(str(cell.value)) if cell.value else 0
                        if len(max_lengths) <= i:
                            max_lengths.append(value_length)
                        else:
                            max_lengths[i] = max(max_lengths[i], value_length)
            else:
                for qual in row['qualifications']:
                    color = "C6EFCE"  # verde chiaro
                    istituto = qual['institute_name'] + " (" + qual['institute_code'] + ")"
                    qualifica = qual['qualification_name']
                    tipo_corso = qual['programme_type_code']
                    nome_corso = qual['degree_course_code']
                    classe = qual['degree_class_code']
                    data_conseg = qual['academic_qualification_date']
                    valutazione = qual['qualification_grade_value']
                    if valutazione == "QUALIFIED" or valutazione == "Abilitato":
                        valutazione = "Abilitato"
                    else:
                        if valutazione == "110L":
                            valutazione = "110 cum laude"
                        valutazione = valutazione + " su " + qual['qualification_grading_scale_maximum_grade']
                    ws.append([cf, istituto, qualifica, tipo_corso, nome_corso, classe, data_conseg, valutazione])
                    
                    fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                    for col in ws.iter_rows(min_row=ws.max_row, max_row=ws.max_row):
                        for i, cell in enumerate(col):
                            cell.fill = fill
                            value_length = len(str(cell.value)) if cell.value else 0
                            if len(max_lengths) <= i:
                                max_lengths.append(value_length)
                            else:
                                max_lengths[i] = max(max_lengths[i], value_length)
        else:
            split_row = row.split()
            cf = split_row[0]
            if "Codice" in split_row[1]:
                color = "FFC7CE"  # rosso
                ws.append([cf, "Codice fiscale non corretto", "N/A", "N/A", "N/A", "N/A", "N/A", "N/A"])
                
                fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                for col in ws.iter_rows(min_row=ws.max_row, max_row=ws.max_row):
                    for i, cell in enumerate(col):
                        cell.fill = fill
                        value_length = len(str(cell.value)) if cell.value else 0
                        if len(max_lengths) <= i:
                            max_lengths.append(value_length)
                        else:
                            max_lengths[i] = max(max_lengths[i], value_length)

    # Imposta larghezza colonne
    for i, width in enumerate(max_lengths, 1):
        col_letter = get_column_letter(i)
        ws.column_dimensions[col_letter].width = width

    # Scrive l'Excel su un buffer
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=EsitoAnisTitoli.xlsx'
    return response


def anist_frequenze_export_excel(request):
    data = request.session.get("multi_data", [])  # Oppure recuperalo come preferisci
    wb = Workbook()
    ws = wb.active
    ws.append(["Codice Fiscale", "Istituto principale", "Plesso", "Tipologia corso", "Anno corso", "Esito frequenza"])

    # Dizionario per la larghezza massima di ogni colonna
    max_lengths = [len(cell.value) for cell in ws[1]]  # Larghezze iniziali dall'intestazione

    if not data:
        return HttpResponse("Nessun dato disponibile per l'esportazione", status=400)

    for i, row in enumerate(data, 1):
        color = "FFFFFF"  # default bianco

        if isinstance(row, list):
            cf = row[0]
            if isinstance(row[1], dict):
                if row[1]['frequentante'] == False :
                    color = "FFC7CE"
                    ws.append([cf, "La richiesta effettuata non produce alcun risultato", "N/A", "N/A", "N/A", "N/A"])
                else:
                    color = "C6EFCE"  # verde chiaro
                    istituto = row[1]['denoIstitutoPrincipale'] + " (" + row[1]['codiceIstitutoPrincipale'] + ")"
                    plesso = row[1]['denominazionePlesso'] + " (" + row[1]['codiceMeccanografico'] + ")"
                    tipo_corso = row[1]['percorsoStudi']
                    anno_accademico = row[1]['annoCorso']
                    esito_map = {1: "Frequentante", 2: "Non Frequentante", 3: "Frequentante su altro anno corso", 4: "Non più Frequentante"}
                    esito = esito_map.get(row[1].get('esitoFrequenza'), row[1].get('esitoFrequenza'))
                    ws.append([cf, istituto, plesso, tipo_corso, anno_accademico, esito])
            else:
                color = "FFC7CE"
                ws.append([cf, row[1], "N/A", "N/A", "N/A", "N/A"])
        else:
            split_row = row.split()
            cf = split_row[0]
            if "Codice" in split_row[1]:
                color = "FFC7CE"  # rosso
                ws.append([cf, "Codice fiscale non corretto", "N/A", "N/A", "N/A", "N/A"])

        fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        for col in ws.iter_rows(min_row=ws.max_row, max_row=ws.max_row):
            for i, cell in enumerate(col):
                cell.fill = fill
                value_length = len(str(cell.value)) if cell.value else 0
                if len(max_lengths) <= i:
                    max_lengths.append(value_length)
                else:
                    max_lengths[i] = max(max_lengths[i], value_length)

    # Imposta larghezza colonne
    for i, width in enumerate(max_lengths, 1):
        col_letter = get_column_letter(i)
        ws.column_dimensions[col_letter].width = width

    # Scrive l'Excel su un buffer
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=EsitoAnistFrequenze.xlsx'
    return response


def anist_frequenze_export_csv(request):
    data = request.session.get("multi_data", [])
    
    response = HttpResponse(
        content_type='text/csv',
        headers={'Content-Disposition': 'attachment; filename="EsitoAnistFrequenze.csv"'},
    )

    writer = csv.writer(response, delimiter=';')
    writer.writerow(["Codice Fiscale", "Istituto principale", "Plesso", "Tipologia corso", "Anno corso", "Esito frequenza"])

    if not data:
        return HttpResponse("Nessun dato disponibile per l'esportazione", status=400)

    for row in data:
        if isinstance(row, list):
            cf = row[0]
            if isinstance(row[1], dict):
                if row[1]['frequentante'] == False:
                    writer.writerow([cf, "La richiesta effettuata non produce alcun risultato", "N/A", "N/A", "N/A", "N/A"])
                else:
                    istituto = row[1]['denoIstitutoPrincipale'] + " (" + row[1]['codiceIstitutoPrincipale'] + ")"
                    plesso = row[1]['denominazionePlesso'] + " (" + row[1]['codiceMeccanografico'] + ")"
                    tipo_corso = row[1]['percorsoStudi']
                    anno_accademico = row[1]['annoCorso']
                    esito_map = {1: "Frequentante", 2: "Non Frequentante", 3: "Frequentante su altro anno corso", 4: "Non più Frequentante"}
                    esito = esito_map.get(row[1].get('esitoFrequenza'), row[1].get('esitoFrequenza'))
                    writer.writerow([cf, istituto, plesso, tipo_corso, anno_accademico, esito])
            else:
                writer.writerow([cf, row[1], "N/A", "N/A", "N/A", "N/A"])
        else:
            split_row = row.split()
            cf = split_row[0]
            if "Codice" in split_row[1]:
                writer.writerow([cf, "Codice fiscale non corretto", "N/A", "N/A", "N/A", "N/A"])

    return response


def anist_titoli_export_excel(request):
    data = request.session.get("multi_data", [])  # Oppure recuperalo come preferisci
    wb = Workbook()
    ws = wb.active
    ws.append(["Codice Fiscale", "Titolo", "Istituto principale", "Plesso", "Votazione"])

    # Dizionario per la larghezza massima di ogni colonna
    max_lengths = [len(cell.value) for cell in ws[1]]  # Larghezze iniziali dall'intestazione

    if not data:
        return HttpResponse("Nessun dato disponibile per l'esportazione", status=400)

    for i, row in enumerate(data, 1):
        color = "FFFFFF"  # default bianco

        if isinstance(row, list):
            cf = row[0]
            if row[1]['presenzaTitoli'] == False:
                color = "FFC7CE"
                valori = [cf, "La richiesta effettuata non produce alcun risultato", "N/A", "N/A", "N/A"]
                ws.append(valori)
            else:
                for titol in row[1]['listaTitoli']:
                    color = "C6EFCE"
                    titolo = titol['denominazioneTitolo']
                    istituto = f"{titol['denoIstitutoPrincipale']} ({titol['codiceIstitutoPrincipale']})"
                    plesso = f"{titol['denominazionePlesso']} ({titol['codiceMeccanografico']})"
                    valutazione = titol['votoFinale']
                    if titol['flagLode'] == "S":
                        valutazione += " con lode"
                    valori = [cf, titolo, istituto, plesso, valutazione]
                    ws.append(valori)
        else:
            split_row = row.split()
            cf = split_row[0]
            if "Codice" in split_row[1]:
                color = "FFC7CE"
                valori = [cf, "Codice fiscale non corretto", "N/A", "N/A", "N/A"]
                ws.append(valori)

        # Applica colore all'ultima riga scritta
        fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        for col in ws.iter_rows(min_row=ws.max_row, max_row=ws.max_row):
            for i, cell in enumerate(col):
                cell.fill = fill
                value_length = len(str(cell.value)) if cell.value else 0
                if len(max_lengths) <= i:
                    max_lengths.append(value_length)
                else:
                    max_lengths[i] = max(max_lengths[i], value_length)

    # Imposta larghezza colonne
    for i, width in enumerate(max_lengths, 1):
        col_letter = get_column_letter(i)
        ws.column_dimensions[col_letter].width = width

    # Scrive l'Excel su un buffer
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=EsitoAnistTitoli.xlsx'
    return response


def anist_titoli_export_csv(request):
    data = request.session.get("multi_data", [])
    
    response = HttpResponse(
        content_type='text/csv',
        headers={'Content-Disposition': 'attachment; filename="EsitoAnistTitoli.csv"'},
    )

    writer = csv.writer(response, delimiter=';')
    writer.writerow(["Codice Fiscale", "Titolo", "Istituto principale", "Plesso", "Votazione"])

    if not data:
        return HttpResponse("Nessun dato disponibile per l'esportazione", status=400)

    for row in data:
        if isinstance(row, list):
            cf = row[0]
            if row[1]['presenzaTitoli'] == False:
                writer.writerow([cf, "La richiesta effettuata non produce alcun risultato", "N/A", "N/A", "N/A"])
            else:
                for titol in row[1]['listaTitoli']:
                    titolo = titol['denominazioneTitolo']
                    istituto = f"{titol['denoIstitutoPrincipale']} ({titol['codiceIstitutoPrincipale']})"
                    plesso = f"{titol['denominazionePlesso']} ({titol['codiceMeccanografico']})"
                    valutazione = titol['votoFinale']
                    if titol['flagLode'] == "S":
                        valutazione += " con lode"
                    writer.writerow([cf, titolo, istituto, plesso, valutazione])
        else:
            split_row = row.split()
            cf = split_row[0]
            if "Codice" in split_row[1]:
                writer.writerow([cf, "Codice fiscale non corretto", "N/A", "N/A", "N/A"])

    return response


def anis_iscrizioni_singola(request):
    if request.user.id:
        utente_sessione = UtentiParametri.objects.get(id=request.user.id)
        utente_abilitato = utente_sessione.anis_IFS02_singolo
        if request.method == 'POST':
            data = []
            cf = request.POST.get('input_CF')
            data.append(cf)
            correttezza_cf = verifica_cf(cf)
            if correttezza_cf == 1 or correttezza_cf == 2:
                data.append(anis_verifica_utente(request.user.username, cf, 1))
                data = converti_data(data)
            else:
                data.append(str(cf) + " Codice fiscale non corretto")

            salva_log(request.user,"Verifica ANIS - IFS02 - Iscrizioni Singolo", "Verificato utente " + cf)
            return render(request, 'anis_iscrizioni_singola.html', {'data': data, 'utente_abilitato': utente_abilitato })
    else:
        utente_abilitato = False

    return render(request, 'anis_iscrizioni_singola.html', { 'utente_abilitato': utente_abilitato })


def anis_iscrizioni_massiva(request):
    if request.user.id:
        utente_sessione = UtentiParametri.objects.get(id=request.user.id)
        utente_abilitato = utente_sessione.anis_IFS02_massivo
        if request.method == 'POST':
            csv_file = request.FILES['cf_csv']
            data = []
            contatore = 0
            if csv_file.name.endswith('.csv'):
                csv_file_text = io.TextIOWrapper(csv_file.file, encoding='utf-8')
                csv_reader = csv.reader(csv_file_text)
                for row in csv_reader:
                    if row[0]:  # Se row[0] non è vuoto o None
                        correttezza_cf = verifica_cf(row[0].strip().upper())
                        if correttezza_cf == 1 or correttezza_cf == 2:
                            data.append(anis_verifica_utente(request.user.username, row[0].strip().upper(), 1))
                            data = converti_data(data)
                        else:
                            data.append( str(row[0].strip().upper()) + " Codice fiscale non corretto")
                        contatore += 1
                salva_log(request.user,"Verifica ANIS - IFS02 - Iscrizioni Massivo", "Verificati n. " + str(contatore) + " CF")
                request.session["multi_data"] = data  # <--- Salva i dati nella sessione
                return render(request, 'anis_iscrizioni_massiva.html', {'data': data, 'utente_abilitato': utente_abilitato })
            elif csv_file.name.endswith('.xlsx') or csv_file.name.endswith('.XLSX') or csv_file.name.endswith('.Xlsx'):
                wb = openpyxl.load_workbook(csv_file)
                sheet = wb.active
                for row in sheet.iter_rows(min_row=1, values_only=True):
                    if row[0]:
                        correttezza_cf = verifica_cf(row[0].strip().upper())
                        if correttezza_cf == 1 or correttezza_cf == 2:
                            data.append(anis_verifica_utente(request.user.username, row[0].strip().upper(), 1))
                            data = converti_data(data)
                        else:
                            data.append( str(row[0].strip().upper()) + " Codice fiscale non corretto")
                        contatore += 1
                salva_log(request.user,"Verifica ANIS - IFS02 - Iscrizioni Massivo", "Verificati n. " + str(contatore) + " CF")
                request.session["multi_data"] = data  # <--- Salva i dati nella sessione
                return render(request, 'anis_iscrizioni_massiva.html', {'data': data, 'utente_abilitato': utente_abilitato })

            else:
                salva_log(request.user,"Verifica ANIS - IFS02 - Iscrizioni Massivo", "Errore caricamento file CSV")
                return render(request, 'anis_iscrizioni_massiva.html', {'error': 'Il file non è un CSV', 'utente_abilitato': utente_abilitato })
    else:
        utente_abilitato = False
    return render(request, 'anis_iscrizioni_massiva.html' , { 'utente_abilitato': utente_abilitato })


def anis_titoli_singola(request):
    if request.user.id:
        utente_sessione = UtentiParametri.objects.get(id=request.user.id)
        utente_abilitato = utente_sessione.anis_IFS03_singolo
        if request.method == 'POST':
            data = []
            cf = request.POST.get('input_CF')
            data.append(cf)
            correttezza_cf = verifica_cf(cf)
            if correttezza_cf == 1 or correttezza_cf == 2:
                data.append(anis_verifica_utente(request.user.username, cf, 2))
                data = converti_data(data)
            else:
                data.append(str(cf) + " Codice fiscale non corretto")

            salva_log(request.user,"Verifica ANIS - IFS03 - Titoli Singolo", "Verificato utente " + cf)
            return render(request, 'anis_titoli_singola.html', {'data': data, 'utente_abilitato': utente_abilitato })
    else:
        utente_abilitato = False
    return render(request, 'anis_titoli_singola.html' , { 'utente_abilitato': utente_abilitato })


def anis_titoli_massiva(request):
    if request.user.id:
        utente_sessione = UtentiParametri.objects.get(id=request.user.id)
        utente_abilitato = utente_sessione.anis_IFS03_massivo
        if request.method == 'POST':
            csv_file = request.FILES['cf_csv']
            data = []
            contatore = 0
            if csv_file.name.endswith('.csv'):
                csv_file_text = io.TextIOWrapper(csv_file.file, encoding='utf-8')
                csv_reader = csv.reader(csv_file_text)
                for row in csv_reader:
                    if row[0]:  # Se row[0] non è vuoto o None
                        correttezza_cf = verifica_cf(row[0].strip().upper())
                        if correttezza_cf == 1 or correttezza_cf == 2:
                            data.append(anis_verifica_utente(request.user.username, row[0].strip().upper(), 2))
                            data = converti_data(data)
                        else:
                            data.append( str(row[0].strip().upper()) + " Codice fiscale non corretto")
                        contatore += 1
                salva_log(request.user,"Verifica ANIS - IFS03 - Titoli Massivo", "Verificati n. " + str(contatore) + " CF")
                # Modifica QUALIFIED in Abilitato
                for item in data:
                    if isinstance(item, dict) and 'qualifications' in item:
                        for qual in item['qualifications']:
                            if qual.get('qualification_grade_value') == 'QUALIFIED':
                                qual['qualification_grade_value'] = 'Abilitato'
                request.session["multi_data"] = data  # <--- Salva i dati nella sessione
                return render(request, 'anis_titoli_massiva.html', {'data': data, 'utente_abilitato': utente_abilitato })

            elif csv_file.name.endswith('.xlsx') or csv_file.name.endswith('.XLSX') or csv_file.name.endswith('.Xlsx'):
                wb = openpyxl.load_workbook(csv_file)
                sheet = wb.active
                for row in sheet.iter_rows(min_row=1, values_only=True):
                    if row[0]:
                        correttezza_cf = verifica_cf(row[0].strip().upper())
                        if correttezza_cf == 1 or correttezza_cf == 2:
                            data.append(anis_verifica_utente(request.user.username, row[0].strip().upper(), 2))
                            data = converti_data(data)
                        else:
                            data.append( str(row[0].strip().upper()) + " Codice fiscale non corretto")
                        contatore += 1
                salva_log(request.user,"Verifica ANIS - IFS03 - Titoli Massivo", "Verificati n. " + str(contatore) + " CF")
                # Modifica QUALIFIED in Abilitato
                for item in data:
                    if isinstance(item, dict) and 'qualifications' in item:
                        for qual in item['qualifications']:
                            if qual.get('qualification_grade_value') == 'QUALIFIED':
                                qual['qualification_grade_value'] = 'Abilitato'
                request.session["multi_data"] = data  # <--- Salva i dati nella sessione
                return render(request, 'anis_titoli_massiva.html', {'data': data, 'utente_abilitato': utente_abilitato })

            else:
                salva_log(request.user,"Verifica ANIS - IFS03 - Titoli Massivo", "Errore caricamento file CSV")
                return render(request, 'anis_titoli_massiva.html', {'error': 'Il file non è un CSV', 'utente_abilitato': utente_abilitato })
    else:
        utente_abilitato = False
    return render(request, 'anis_titoli_massiva.html' , { 'utente_abilitato': utente_abilitato })



def anist_frequenze_singola(request):
    if request.user.id:
        utente_sessione = UtentiParametri.objects.get(id=request.user.id)
        utente_abilitato = utente_sessione.anist_frequenze_singolo
        if request.method == 'POST':
            data = []
            cf = request.POST.get('input_CF')
            data.append(cf)
            correttezza_cf = verifica_cf(cf)
            if correttezza_cf == 1 or correttezza_cf == 2:
                data.append(anis_verifica_utente(request.user.username, cf, 3))
                data = converti_data(data)
            else:
                data.append(str(cf) + " Codice fiscale non corretto")

            salva_log(request.user,"Verifica ANIST - Frequenze Singolo", "Verificato utente " + cf)
            return render(request, 'anist_frequenze_singola.html', {'data': data, 'utente_abilitato': utente_abilitato })
    else:
        utente_abilitato = False

    return render(request, 'anist_frequenze_singola.html', { 'utente_abilitato': utente_abilitato })


def anist_frequenze_massiva(request):
    if request.user.id:
        utente_sessione = UtentiParametri.objects.get(id=request.user.id)
        utente_abilitato = utente_sessione.anist_frequenze_massivo
        if request.method == 'POST':
            csv_file = request.FILES['cf_csv']
            data = []
            contatore = 0
            if csv_file.name.endswith('.csv'):
                csv_file_text = io.TextIOWrapper(csv_file.file, encoding='utf-8')
                csv_reader = csv.reader(csv_file_text)
                for row in csv_reader:
                    if row[0]:  # Se row[0] non è vuoto o None
                        correttezza_cf = verifica_cf(row[0].strip().upper())
                        if correttezza_cf == 1 or correttezza_cf == 2:
                            data.append((row[0].strip().upper(), anis_verifica_utente(request.user.username, row[0].strip().upper(), 3)))
                            data = converti_data(data)
                        else:
                            data.append( str(row[0].strip().upper()) + " Codice fiscale non corretto")
                        contatore += 1
                salva_log(request.user,"Verifica ANIST - Frequenze Massivo", "Verificati n. " + str(contatore) + " CF")
                request.session["multi_data"] = data  # <--- Salva i dati nella sessione
                return render(request, 'anist_frequenze_massiva.html', {'data': data, 'utente_abilitato': utente_abilitato })
            elif csv_file.name.endswith('.xlsx') or csv_file.name.endswith('.XLSX') or csv_file.name.endswith('.Xlsx'):
                wb = openpyxl.load_workbook(csv_file)
                sheet = wb.active
                for row in sheet.iter_rows(min_row=1, values_only=True):
                    if row[0]:
                        correttezza_cf = verifica_cf(row[0].strip().upper())
                        if correttezza_cf == 1 or correttezza_cf == 2:
                            data.append((row[0].strip().upper(), anis_verifica_utente(request.user.username, row[0].strip().upper(), 3)))
                            data = converti_data(data)
                        else:
                            data.append(str(row[0].strip().upper()) + " Codice fiscale non corretto")
                        contatore += 1
                salva_log(request.user,"Verifica ANIST - Frequenze Massivo", "Verificati n. " + str(contatore) + " CF")
                request.session["multi_data"] = data  # <--- Salva i dati nella sessione
                return render(request, 'anist_frequenze_massiva.html', {'data': data, 'utente_abilitato': utente_abilitato })
            else:
                salva_log(request.user,"Verifica ANIST - Frequenze Massivo", "Errore caricamento file CSV")
                return render(request, 'anist_frequenze_massiva.html', {'error': 'Il file non è un CSV', 'utente_abilitato': utente_abilitato })
    else:
        utente_abilitato = False
    return render(request, 'anist_frequenze_massiva.html', { 'utente_abilitato': utente_abilitato })


def anist_titoli_singola(request):
    if request.user.id:
        utente_sessione = UtentiParametri.objects.get(id=request.user.id)
        utente_abilitato = utente_sessione.anist_titoli_singolo
        if request.method == 'POST':
            data = []
            cf = request.POST.get('input_CF')
            data.append(cf)
            correttezza_cf = verifica_cf(cf)
            if correttezza_cf == 1 or correttezza_cf == 2:
                data.append(anis_verifica_utente(request.user.username, cf, 4))
                data = converti_data(data)
            else:
                data.append(str(cf) + " Codice fiscale non corretto")

            salva_log(request.user,"Verifica ANIST - Titoli Singolo", "Verificato utente " + cf)
            return render(request, 'anist_titoli_singola.html', {'data': data, 'utente_abilitato': utente_abilitato })
    else:
        utente_abilitato = False

    return render(request, 'anist_titoli_singola.html', { 'utente_abilitato': utente_abilitato })


def anist_titoli_massiva(request):
    if request.user.id:
        utente_sessione = UtentiParametri.objects.get(id=request.user.id)
        utente_abilitato = utente_sessione.anist_titoli_massivo
        if request.method == 'POST':
            csv_file = request.FILES['cf_csv']
            data = []
            contatore = 0
            if csv_file.name.endswith('.csv'):
                csv_file_text = io.TextIOWrapper(csv_file.file, encoding='utf-8')
                csv_reader = csv.reader(csv_file_text)
                for row in csv_reader:
                    if row[0]:  # Se row[0] non è vuoto o None
                        correttezza_cf = verifica_cf(row[0].strip().upper())
                        if correttezza_cf == 1 or correttezza_cf == 2:
                            data.append((row[0].strip().upper(), anis_verifica_utente(request.user.username, row[0].strip().upper(), 4)))
                            data = converti_data(data)
                        else:
                            data.append( str(row[0].strip().upper()) + " Codice fiscale non corretto")
                        contatore += 1
                salva_log(request.user,"Verifica ANIST - Titoli Massivo", "Verificati n. " + str(contatore) + " CF")
                request.session["multi_data"] = data  # <--- Salva i dati nella sessione
                return render(request, 'anist_titoli_massiva.html', {'data': data, 'utente_abilitato': utente_abilitato })
            elif csv_file.name.endswith('.xlsx') or csv_file.name.endswith('.XLSX') or csv_file.name.endswith('.Xlsx'):
                wb = openpyxl.load_workbook(csv_file)
                sheet = wb.active
                for row in sheet.iter_rows(min_row=1, values_only=True):
                    if row[0]:
                        correttezza_cf = verifica_cf(row[0].strip().upper())
                        if correttezza_cf == 1 or correttezza_cf == 2:
                            data.append((row[0].strip().upper(), anis_verifica_utente(request.user.username, row[0].strip().upper(), 4)))
                            data = converti_data(data)
                        else:
                            data.append(str(row[0].strip().upper()) + " Codice fiscale non corretto")
                        contatore += 1
                salva_log(request.user,"Verifica ANIST - Titoli Massivo", "Verificati n. " + str(contatore) + " CF")
                request.session["multi_data"] = data  # <--- Salva i dati nella sessione
                return render(request, 'anist_titoli_massiva.html', {'data': data, 'utente_abilitato': utente_abilitato })
            else:
                salva_log(request.user,"Verifica ANIST - Titoli Massivo", "Errore caricamento file CSV")
                return render(request, 'anist_titoli_massiva.html', {'error': 'Il file non è un CSV', 'utente_abilitato': utente_abilitato })
    else:
        utente_abilitato = False
    return render(request, 'anist_titoli_massiva.html', { 'utente_abilitato': utente_abilitato })


def anis_verifica_utente(user_ID, cf, id_caso):
    parametri_anis = AnisParametri.objects.get(id=id_caso)
    caso = ''
    if id_caso == 1:
        caso = "IFS02"
    if id_caso == 2:
        caso = "IFS03"
    if id_caso == 3:
        caso = "Anist_Frequenze"
    else:
        caso = "Anist_Titoli"

    kid = parametri_anis.kid
    alg = parametri_anis.alg
    typ = parametri_anis.typ
    issuer = parametri_anis.iss
    subject = parametri_anis.sub
    aud = parametri_anis.aud
    purposeid = parametri_anis.purposeid
    audience = parametri_anis.audience
    baseurlauth = parametri_anis.baseurlauth
    target = parametri_anis.target
    clientid = parametri_anis.clientid
    private_key = parametri_anis.private_key
    userid = user_ID
    location = 'PortaleOpenMSP'
    loa = 'LoA2'
    if id_caso == 1 or id_caso == 2:
        richiesta = f'{{"tax_code":"{cf}"}}'
    else:
        richiesta = f'{{"codiceFiscale":"{cf}"}}'


    issued = datetime.datetime.utcnow()
    delta = datetime.timedelta(minutes=43200)
    expire_in = issued + delta
    dnonce = random.randint(1000000000000, 9999999999999)

    headers_rsa = {
        "kid": kid,
        "alg": alg,
        "typ": typ
    }

    jti = uuid.uuid4()
    audit_payload = {
        "userID": userid,
        "userLocation": location,
        "LoA": loa,
        "iss" : clientid,
        "aud" : audience,
        "purposeId": purposeid,
        "dnonce" : dnonce,
        "jti":str(jti),
        "iat": issued,
        "nbf" : issued,
        "exp": expire_in
        }

    audit = jwt.encode(audit_payload, private_key, algorithm=Algorithms.RS256, headers=headers_rsa)
    audit_hash = hashlib.sha256(audit.encode('UTF-8')).hexdigest()

    jti = uuid.uuid4()
    payload = {
        "iss": clientid,
        "sub": clientid,
        "aud": aud,
        "purposeId": purposeid,
        "jti": str(jti),
        "iat": issued,
        "exp": expire_in,
        "digest": {
            "alg": "SHA256",
            "value": audit_hash
            }
        }

    client_assertion = jwt.encode(payload, private_key, algorithm=Algorithms.RS256, headers=headers_rsa)

    params = urllib.parse.urlencode({
        'client_id': clientid,
        'client_assertion': client_assertion,
        'client_assertion_type': 'urn:ietf:params:oauth:client-assertion-type:jwt-bearer',
        'grant_type': 'client_credentials'
        })

    headers = {"Content-type": "application/x-www-form-urlencoded"}
    conn = http.client.HTTPSConnection(re.sub(r'^https?://', '', baseurlauth))
    conn.request("POST", "/token.oauth2", params, headers)
    response = conn.getresponse()
    voucher = json.loads(response.read())["access_token"]

    # prepara il body per la richiesta e relativo digest
    body = richiesta
    type = 'application/json'
    encoding = 'UTF-8'
    body_digest = hashlib.sha256(body.encode('UTF-8'))
    digest = 'SHA-256=' + base64.b64encode(body_digest.digest()).decode('UTF-8')

    # crea signature
    jti = uuid.uuid4()
    payload = {
        "iss" : clientid,
        "aud" : audience,
        "purposeId": purposeid,
        "sub": clientid,
        "jti": str(jti),
        "iat": issued,
        "nbf" : issued,
        "exp": expire_in,
        "signed_headers": [
            {"digest": digest},
            {"content-type": type},
            {"content-encoding": encoding}
            ]
        }
    signature = jwt.encode(payload, private_key, algorithm=Algorithms.RS256, headers=headers_rsa)
    # effettua chiamata
    api_url = target
    headers =  {"Accept":"application/json",
                "Content-Type":type,
                "Content-Encoding":encoding,
                "Digest":digest,
                "Authorization":"Bearer " + voucher,
                "Agid-JWT-TrackingEvidence":audit,
                "Agid-JWT-Signature":signature
                }

    response = requests.post(api_url, data=body.encode('UTF-8'), headers=headers, verify=False)
    if response.status_code == 403:
        return str(cf) + "Errore di comunicazione con la API"
    elif response.status_code == 500:
        return str(cf) + " La richiesta effettuata non produce alcun risultato"
    else:
        return response.json()

def impostazioni_anis(request):
    servizi_anis = AnisServizi.objects.all()
    parametri_anis = AnisParametri.objects.all()

    service_active = ServiziParametri.objects.all()
    i_serv=0
    service_desc = ["" for _ in range(ServiziParametri.objects.count())]
    for servizio in service_active:
        service_desc[i_serv]= (servizio.attivo)
        i_serv += 1

    if request.method == 'POST':
        posizione_servizio = ServiziParametri.objects.filter(gruppo_id=6).values_list('id', flat=True)
        for i in range(1, AnisParametri.objects.count()+1):
            if service_desc[posizione_servizio[i-1]-1]:
                kid = request.POST.get('kid' + str(i))
                alg = request.POST.get('alg' + str(i))
                typ = request.POST.get('typ' + str(i))
                iss = request.POST.get('iss' + str(i))
                sub = request.POST.get('sub' + str(i))
                aud = request.POST.get('aud' + str(i))
                purposeid = request.POST.get('purposeid' + str(i))
                audience = request.POST.get('audience' + str(i))
                baseurlauth = request.POST.get('baseurlauth' + str(i))
                target = request.POST.get('target' + str(i))
                clientid = request.POST.get('clientid' + str(i))
                private_key = request.POST.get('private_key' + str(i))
                ver_eservice = request.POST.get('ver_eservice' + str(i))
                dati = AnisParametri(i, i, kid, alg, typ, iss, sub, aud, purposeid, audience, baseurlauth, target, clientid, private_key, ver_eservice)
                dati.save()
        salva_log(request.user,"Impostazioni ANIS", "modifica parametri")

    return render(request, 'impostazioni_anis.html', { 'servizi_anis': servizi_anis, 'parametri_anis': parametri_anis })

