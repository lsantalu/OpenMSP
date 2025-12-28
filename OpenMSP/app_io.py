from django.http import HttpResponse
from django.contrib.auth.models import User
from django.shortcuts import render
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from markdown_it import MarkdownIt
import datetime

from impostazioni.models import AppIoParametri
from impostazioni.models import AppIoCatalogoServizi
from impostazioni.models import AppIoCatalogoArgomenti
from impostazioni.models import AppIoElencoMessaggi

from impostazioni.models import UtentiParametri

from .utils import converti_data
from .utils import normalizza_data
from .utils import is_daylight_saving
from .utils import salva_log

from .verifica_cf import verifica_cf
import requests
import io
import csv
import json
import openpyxl
from markdown_it import MarkdownIt
import html2text
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment
from openpyxl.utils import get_column_letter



def app_io_storico_pagina_export_excel(request):
    messaggi = request.session.get("multi_messaggi", [])  # Oppure recuperalo come preferisci
    page_number = request.GET.get('page', 1)

    elementi_per_pagina = 20

    paginator = Paginator(messaggi, elementi_per_pagina)
    try:
        page_obj = paginator.page(page_number)
    except PageNotAnInteger:
        page_obj = paginator.page(1)
    except EmptyPage:
        page_obj = paginator.page(paginator.num_pages)


    wb = Workbook()
    ws = wb.active
    ws.append(["ID Msg", "Servizio", "Titolo messaggio", "CF Destinatario", "Data invio", "Utente mitt.", "Esito"])

    # Dizionario per la larghezza massima di ogni colonna
    max_lengths = [len(cell.value) for cell in ws[1]]  # Larghezze iniziali dall'intestazione

    if not messaggi:
        return HttpResponse("Nessun dato disponibile per l'esportazione", status=400)

    for row in page_obj.object_list:
        if "Errore" in row['esito']:
            color = "FFC7CE"
            stato = "Errore"
        else:
            color = "C6EFCE"  # rosso se contiene "Errore", altrimenti verde chiaro
            stato = "Successo"

        nome_utente = User.objects.get(id=row['utente_id']).username ##servizi_utente = UtentiParametri.objects.get(id=utente_id)
        servizio = AppIoCatalogoServizi.objects.get(id=row['servizio_id']).servizio

        fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

        ws.append([row['id'], servizio, row['titolo'], row['cf_destinatario'], row['timestamp'], nome_utente, stato])
        for col in ws.iter_rows(min_row=ws.max_row, max_row=ws.max_row):
            for i, cell in enumerate(col):
                cell.fill = fill
                value_length = len(str(cell.value)) if cell.value else 0
                if len(max_lengths) <= i:
                    max_lengths.append(value_length)
                else:
                    max_lengths[i] = max(max_lengths[i], value_length)

    # Imposta larghezza colonne
    for i, width in enumerate(max_lengths, 1):
        col_letter = get_column_letter(i)
        ws.column_dimensions[col_letter].width = width

    # Scrive l'Excel su un buffer
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=ElencoMessaggiAppIO.xlsx'
    return response


def app_io_storico_full_export_excel(request):
    messaggi = request.session.get("multi_messaggi", [])  # Oppure recuperalo come preferisci

    wb = Workbook()
    ws = wb.active
    ws.append(["ID Msg", "Servizio", "Titolo messaggio", "CF Destinatario", "Data invio", "Utente mitt.", "Esito"])

    # Dizionario per la larghezza massima di ogni colonna
    max_lengths = [len(cell.value) for cell in ws[1]]  # Larghezze iniziali dall'intestazione

    if not messaggi:
        return HttpResponse("Nessun dato disponibile per l'esportazione", status=400)

    for i, row in enumerate(messaggi, 1):
        color = "FFC7CE" if "Errore" in row['esito'] else "C6EFCE"  # rosso se contiene "Errore", altrimenti verde chiaro
        stato = "Errore" if "Errore" in row['esito'] else "Successo"

        nome_utente = User.objects.get(id=row['utente_id']).username ##servizi_utente = UtentiParametri.objects.get(id=utente_id)
        servizio = AppIoCatalogoServizi.objects.get(id=row['servizio_id']).servizio

        fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

        ws.append([row['id'], servizio, row['titolo'], row['cf_destinatario'], row['timestamp'], nome_utente, stato])
        for col in ws.iter_rows(min_row=ws.max_row, max_row=ws.max_row):
            for i, cell in enumerate(col):
                cell.fill = fill
                value_length = len(str(cell.value)) if cell.value else 0
                if len(max_lengths) <= i:
                    max_lengths.append(value_length)
                else:
                    max_lengths[i] = max(max_lengths[i], value_length)

    # Imposta larghezza colonne
    for i, width in enumerate(max_lengths, 1):
        col_letter = get_column_letter(i)
        ws.column_dimensions[col_letter].width = width

    # Scrive l'Excel su un buffer
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=ElencoFullMessaggiAppIO.xlsx'
    return response


def app_io_composer_export_excel(request):
    data = request.session.get("multi_data", [])  # Oppure recuperalo come preferisci
    wb = Workbook()
    ws = wb.active
    ws.append(["Servizio", "Titolo messaggio", "CF Destinatario", "Data invio", "Utente mitt.", "Esito"])

    # Dizionario per la larghezza massima di ogni colonna
    max_lengths = [len(cell.value) for cell in ws[1]]  # Larghezze iniziali dall'intestazione

    if not data:
        return HttpResponse("Nessun dato disponibile per l'esportazione", status=400)

    for i, row in enumerate(data, 1):
        stato = row[2]
        color = "FFFFFF"  # default bianco

        if stato == "0":
            color = "FFC7CE"  # giallo
        else :
            color = "C6EFCE"  # verde chiaro

        fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

        ws.append([i, row[0], row[1], row[2], row[3]])
        for col in ws.iter_rows(min_row=ws.max_row, max_row=ws.max_row):
            for i, cell in enumerate(col):
                cell.fill = fill
                value_length = len(str(cell.value)) if cell.value else 0
                if len(max_lengths) <= i:
                    max_lengths.append(value_length)
                else:
                    max_lengths[i] = max(max_lengths[i], value_length)

    # Imposta larghezza colonne
    for i, width in enumerate(max_lengths, 1):
        col_letter = get_column_letter(i)
        ws.column_dimensions[col_letter].width = width

    # Scrive l'Excel su un buffer
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=EsitoAppIoComposer.xlsx'
    return response


def app_io_massivo_export_excel(request):
    data = request.session.get("multi_data", [])  # Oppure recuperalo come preferisci
    wb = Workbook()
    ws = wb.active
    ws.append(["#", "CF", "Verifica CF", "Esito invio", "ID/Errore"])

    # Dizionario per la larghezza massima di ogni colonna
    max_lengths = [len(cell.value) for cell in ws[1]]  # Larghezze iniziali dall'intestazione

    if not data:
        return HttpResponse("Nessun dato disponibile per l'esportazione", status=400)

    for i, row in enumerate(data, 1):
        stato = row[2]
        color = "FFFFFF"  # default bianco

        if stato == "0":
            color = "FFC7CE"  # giallo
        else :
            color = "C6EFCE"  # verde chiaro

        fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

        ws.append([i, row[0], row[1], row[2], row[3]])
        for col in ws.iter_rows(min_row=ws.max_row, max_row=ws.max_row):
            for i, cell in enumerate(col):
                cell.fill = fill
                value_length = len(str(cell.value)) if cell.value else 0
                if len(max_lengths) <= i:
                    max_lengths.append(value_length)
                else:
                    max_lengths[i] = max(max_lengths[i], value_length)

    # Imposta larghezza colonne
    for i, width in enumerate(max_lengths, 1):
        col_letter = get_column_letter(i)
        ws.column_dimensions[col_letter].width = width

    # Scrive l'Excel su un buffer
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=EsitoAppIoMassiviPagina.xlsx'
    return response


def app_io_verifica_massivo_export_excel(request):
    data = request.session.get("multi_data", [])  # Oppure recuperalo come preferisci
    wb = Workbook()
    ws = wb.active
    ws.append(["#", "CF", "Verifica", "Stato App IO"])

    # Dizionario per la larghezza massima di ogni colonna
    max_lengths = [len(cell.value) for cell in ws[1]]  # Larghezze iniziali dall'intestazione

    if not data:
        return HttpResponse("Nessun dato disponibile per l'esportazione", status=400)

    for i, row in enumerate(data, 1):
        split_row = row.split()
        stato = split_row[1]
        color = "FFFFFF"  # default bianco
        desc_stato = ""

        if stato == "0" in split_row[1]:
            color = "FFEB9C"  # giallo
            desc_stato = "Utente attivo ma servizio non attivo"
        elif stato == "1" in split_row[1]:
            color = "C6EFCE"  # verde chiaro
            desc_stato = "Servizio attivo"
        elif stato == "2":
            color = "FFEB9C"  # giallo
            desc_stato = "Persona minorenne"
        elif stato == "7":
            color = "FFC7CE"  # rosso
            desc_stato = "Utente non attivo"
        elif stato == "8":
            color = "FFC7CE"  # rosso
            desc_stato = "Utente non trovato"
        elif stato == "-1":
            color = "FFC7CE"  # rosso
            desc_stato = "Codice fiscale errato"
        else:
            color = "FFC7CE"  # rosso
            desc_stato = "Errore di comunicazion"

        fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

        ws.append([i, split_row[0], split_row[1], desc_stato])
        for col in ws.iter_rows(min_row=ws.max_row, max_row=ws.max_row):
            for i, cell in enumerate(col):
                cell.fill = fill
                value_length = len(str(cell.value)) if cell.value else 0
                if len(max_lengths) <= i:
                    max_lengths.append(value_length)
                else:
                    max_lengths[i] = max(max_lengths[i], value_length)

    # Imposta larghezza colonne
    for i, width in enumerate(max_lengths, 1):
        col_letter = get_column_letter(i)
        ws.column_dimensions[col_letter].width = width

    # Scrive l'Excel su un buffer
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=EsitoAppIoMassiviFull.xlsx'
    return response


def app_io_esempio_markdown(request):
    return render(request, 'app_io_esempio_markdown.html')


def app_io_invio_messaggio(url, cf, chiave_servizio, titolo, messaggio, dataScadenza, input_IUV, mezzo1, testobottone1, comandobottone1, mezzo2, testobottone2, comandobottone2):

    if not 10 < len(titolo) < 120:
        return HttpResponse('91', status=200)
    if not 80 < len(messaggio) < 10000:
        return HttpResponse('92', status=200)

    if comandobottone1 is not None :
        if isinstance(comandobottone1, (int, float)):
            comandobottone1 = str(comandobottone1)
        if comandobottone1.isdigit() :
            if not comandobottone1.startswith("+"):
                comandobottone1 = "+" + comandobottone1

    if comandobottone1 is not None and comandobottone2 is not None :
        if isinstance(comandobottone2, (int, float)):
            comandobottone2 = str(comandobottone2)
        if comandobottone2.isdigit() :
            if not comandobottone2.startswith("+"):
                comandobottone2 = "+" + comandobottone2

    messaggio_CTA = ""
    messaggio_CTA1 = ""
    messaggio_CTA2 = ""
    action1 = ""
    action2 = ""
    if mezzo1:
        if mezzo1 == "web1" and comandobottone1 != "" :
            action1 = "iohandledlink://"+comandobottone1
        elif mezzo1 == "mail1" and comandobottone1 != "" :
            action1 = "iohandledlink://mailto:"+comandobottone1
        elif mezzo1 == "sms1" and comandobottone1 != "" :
            action1 = "iohandledlink://sms:"+comandobottone1
        elif mezzo1 == "tel1" and comandobottone1 != "" :
            action1 = "iohandledlink://tel:"+comandobottone1

        if action1 == "" or testobottone1 == "" :
            messaggio_CTA1 =""
        else:
            messaggio_CTA1 = (
                "---\nit:\n"
                " cta_1:\n"
                f"  text: \"{testobottone1}\"\n  action: \"iohandledlink://{action1}\"\n"
            )
        if mezzo1 and mezzo2:
            if mezzo2 == "web2" and comandobottone2 != "" :
                action2 = "iohandledlink://"+comandobottone2
            elif mezzo2 == "mail2" and comandobottone2 != "" :
                action2 = "iohandledlink://mailto:"+comandobottone2
            elif mezzo2 == "sms2" and comandobottone2 != "" :
                action2 = "iohandledlink://sms:"+comandobottone2
            elif mezzo2 == "tel2" and comandobottone2 != "" :
                action2 = "iohandledlink://tel:"+comandobottone2

            if action2 == "" or testobottone2 == "" :
                messaggio_CTA2 =""
            else:
                messaggio_CTA2 = (
                " cta_2: \n"
                f"  text: \"{testobottone2}\"\n  action: \"{action2}\"\n"
            )
        if messaggio_CTA1 != "" :
            messaggio_CTA = messaggio_CTA1 + messaggio_CTA2 + "---\n\n"
            messaggio = messaggio_CTA + messaggio

    headers = {
        'Content-Type': 'application/json',
        'Ocp-Apim-Subscription-Key': chiave_servizio
    }

    dataScadenzaPulita = ""
    if dataScadenza != None and dataScadenza != "":
        dataScadenzaAux = normalizza_data(dataScadenza)
        if isinstance(dataScadenzaAux, datetime.datetime):
            oggi = datetime.datetime.today()
        else:
            oggi = datetime.date.today()
        if dataScadenzaAux < oggi:
            dataScadenzaAux = oggi + datetime.timedelta(days=365)

        timezone = 'Europe/Rome'
        if is_daylight_saving(dataScadenzaAux, timezone):
            dataScadenzaPulita = dataScadenzaAux.strftime('%Y-%m-%d') + "T21:59:59.000Z"
        else:
            dataScadenzaPulita = dataScadenzaAux.strftime('%Y-%m-%d') + "T22:59:59.000Z"

    if input_IUV:
        importoPulito = 1
        input_IUV_pulito = input_IUV.replace(" ", "").replace("'", "")

        data = {
            "time_to_live": 604800,
            "content": {
                "subject": titolo,
                "markdown": messaggio,
                "payment_data": {
                    "amount": importoPulito,
                    "notice_number": input_IUV_pulito,
                    "invalid_after_due_date": True
                },
            **({"due_date": dataScadenzaPulita} if dataScadenza != "" else {})
            },
            "fiscal_code": cf,
            "feature_level_type": "STANDARD"
        }
    else:
        data = {
            "time_to_live": 604800,
            "content": {
                "subject": titolo,
                "markdown": messaggio,
                **({"due_date": dataScadenzaPulita} if dataScadenza != None and dataScadenza != "" else {})
            },
            "fiscal_code": cf,
            "feature_level_type": "STANDARD"
        }

    return requests.post(url, headers=headers, json=data)


def app_io_singolo(request):
    voci_catalogo = AppIoCatalogoArgomenti.objects.all()
    voci_servizio = AppIoCatalogoServizi.objects.all()
    app_io_parametri = AppIoParametri.objects.get(id=1)
    url = app_io_parametri.api_url + '/messages'
    if request.user.id:
        utente_sessione = UtentiParametri.objects.get(id=request.user.id)
        utente_abilitato = utente_sessione.app_io_singolo
        if request.method == 'POST':
            cf = request.POST.get('input_CF').strip().upper()
            id_servizio = request.POST.get('sceltaServizio')
            servizioScelto = AppIoCatalogoServizi.objects.get(id=id_servizio)
            titolo = request.POST.get('subject')
            messaggio = request.POST.get('MessageArea')
            toggleScadenza = request.POST.get('toggleScadenza')
            dataScadenza = request.POST.get('dataScadenza')
            dataScadenzaItalia = ""
            if dataScadenza != "" :
                dataScadenzaItalia = converti_data(dataScadenza)
            togglePagamento = request.POST.get('togglePagamento')
            input_IUV = request.POST.get('input_IUV')
            conferma = request.POST.get('conferma')
            bottone1 = request.POST.get('toggleBottone1')
            mezzo1 = request.POST.get('radio_bottone1')
            testobottone1 = request.POST.get('testo_bottone1')
            comandobottone1 = request.POST.get('comando_bottone1')
            bottone2 = request.POST.get('toggleBottone2')#13
            mezzo2 = request.POST.get('radio_bottone2')
            testobottone2 = request.POST.get('testo_bottone2')
            comandobottone2 = request.POST.get('comando_bottone2')
            data = []
            if conferma: ##ricevuto da conferma.html
                correttezza_cf = verifica_cf(cf)
                if  correttezza_cf == 2: #CODICE FISCALE MINORENNE
                    data = "CF di persona minorenne"
                elif correttezza_cf == -1: ##CODICE FISCALE SBAGLIATO
                    data = "Codice fiscale errato"
                else: ##CODICE FISCALE CORRETTO
                    risposta = app_io_verifica_utente_attivo(cf, servizioScelto.chiave_api)
                    if  risposta == 1: ##UTENTE ATTIVO
                        response = app_io_invio_messaggio(url, cf, servizioScelto.chiave_api, titolo, messaggio , dataScadenza, input_IUV, mezzo1, testobottone1, comandobottone1, mezzo2, testobottone2, comandobottone2)
                        if response.status_code == 201:
                            response_json = response.json()
                            id_messaggio = response_json.get("id")
                            data = (
    f"Messaggio inviato con successo a <b>{cf}</b><br><br>"
    f"ID messaggio = <b>{id_messaggio}</b>"
)
                        else:
                            data = "Errore di comunicazione"
                            id_messaggio = data
                    elif risposta == 0: ##UTENTE ATTIVO NON PERMESSO
                        data = "Utente attivo ma servizio non attivo"
                        id_messaggio = data
                    elif risposta == 7: ##UTENTE NON ATTIVO
                        data = "Utente non attivo"
                        id_messaggio = data
                    elif risposta == 8: ##UTENTE NON TROVATO
                        data = "Utente non trovato"
                        id_messaggio = data
                    else:
                        data = "Errore di comunicazione"
                        id_messaggio = data

                    app_io_salva_messaggio(request.user, servizioScelto, cf, titolo, messaggio, dataScadenza, input_IUV, mezzo1, testobottone1, comandobottone1, mezzo2, testobottone2, comandobottone2, id_messaggio)
                salva_log(request.user,"App IO singolo", "Invio messaggio del servizio " + servizioScelto.servizio + " a " + cf)
                return render(request, 'app_io_singolo.html', {'data': data, 'utente_abilitato': utente_abilitato })
            else:
                data.append(cf) #--0
                data.append(id_servizio) #--1
                data.append(titolo) #--2
                data.append(messaggio) #--3
                md = MarkdownIt()
                data.append(md.render(messaggio)) #--4
                data.append(toggleScadenza) #--5
                data.append(dataScadenza) #--6
                data.append(dataScadenzaItalia) #--7
                data.append(togglePagamento) #--8
                data.append(input_IUV) #--9
                data.append(bottone1) #--10
                data.append(mezzo1) #--11
                data.append(testobottone1) #--12
                data.append(comandobottone1) #--13
                data.append(bottone2) #--14
                data.append(mezzo2) #--15
                data.append(testobottone2) #--16
                data.append(comandobottone2) #--17

                return render(request, 'app_io_singolo_conferma_prev.html', {'dati_catalogo': voci_catalogo, 'dati_servizio': voci_servizio, 'data': data, 'utente_abilitato': utente_abilitato })
    else:
        utente_abilitato = False
    return render(request, 'app_io_singolo.html', { 'dati_catalogo': voci_catalogo, 'dati_servizio': voci_servizio, 'utente_abilitato': utente_abilitato })


def app_io_massivo(request):
    voci_catalogo = AppIoCatalogoArgomenti.objects.all()
    voci_servizio = AppIoCatalogoServizi.objects.all()
    app_io_parametri = AppIoParametri.objects.get(id=1)
    url = app_io_parametri.api_url + '/messages'
    if request.user.id:
        utente_sessione = UtentiParametri.objects.get(id=request.user.id)
        utente_abilitato = utente_sessione.app_io_massivo
        if request.method == 'POST':
            csv_file = request.FILES['cf_csv']
            id_servizio = request.POST.get('sceltaServizio')
            servizioScelto = AppIoCatalogoServizi.objects.get(id=id_servizio)
            data = []
            contatore = 0
            if csv_file.name.endswith('.csv') or csv_file.name.endswith('.CSV') or csv_file.name.endswith('.Csv'):
                csv_file_text = io.TextIOWrapper(csv_file.file, encoding='utf-8')
                csv_reader = csv.reader(csv_file_text)
                next(csv_reader)
                for row in csv_reader:
                    ###if row and (len(row) == 3 or len(row) == 6):  # Se row[0] non è vuoto o None
                    if row[0]:  # Se row[0] non è vuoto o None
                        cf = row[0].strip().upper()
                        titolo = row[1]
                        messaggio = row[2]
                        dataScadenza = ""
                        input_IUV = ""
                        mezzo1 = ""
                        testobottone1 = ""
                        comandobottone1 = ""
                        mezzo2 = ""
                        testobottone2 = ""
                        comandobottone2 = ""
                        if len(row) > 3 and row[3]:
                            dataScadenza = row[3]
                        if len(row) > 4 and row[4]:
                            input_IUV = row[4]
                        if len(row) > 5 and row[5]:
                            mezzo1 = row[5]+"1"
                        if len(row) > 6 and row[6]:
                            testobottone1 = row[6]
                        if len(row) > 7 and row[7]:
                            comandobottone1 = row[7]
                        if len(row) > 8 and row[8]:
                            mezzo2 = row[8]+"2"
                        if len(row) > 9 and row[9]:
                            testobottone2 = row[9]
                        if len(row) > 10 and row[10]:
                            comandobottone2 = row[10]

                        correttezza_cf = verifica_cf(cf)
                        if  correttezza_cf == 2: #CODICE FISCALE MINORENNE
                            data.append((row[0].strip().upper(), correttezza_cf , correttezza_cf, "CF di persona minorenne"))
                        elif correttezza_cf == -1: ##CODICE FISCALE SBAGLIATO
                            data.append((row[0].strip().upper(), correttezza_cf , correttezza_cf, "Codice fiscale errato"))
                        else:
                            risposta = app_io_verifica_utente_attivo(cf, servizioScelto.chiave_api)
                            if  risposta == 1: ##UTENTE ATTIVO
                                response = app_io_invio_messaggio(url, cf, servizioScelto.chiave_api, titolo, messaggio, dataScadenza, input_IUV, mezzo1, testobottone1, comandobottone1, mezzo2, testobottone2, comandobottone2)
                                if response.status_code == 201:
                                    response_json = response.json()
                                    id_messaggio = response_json.get("id")
                                    testo = "ID: " + id_messaggio
                                elif response.content.decode() == '91':
                                    testo = "Titolo troppo corto"
                                    id_messaggio = testo
                                    risposta = 9
                                elif response.content.decode() == '92':
                                    testo = "Messaggio troppo corto"
                                    id_messaggio = testo
                                    risposta = 9
                                else:
                                    testo = "Errore di comunicazione"
                                    id_messaggio = testo
                            elif risposta == 0: ##UTENTE ATTIVO NON PERMESSO
                                testo = "Utente attivo ma servizio non attivo"
                                id_messaggio = testo
                            elif risposta == 7: ##UTENTE NON ATTIVO
                                testo = "Utente non attivo"
                                id_messaggio = testo
                            elif risposta == 8: ##UTENTE NON TROVATO
                                testo = "Utente non trovato"
                                id_messaggio = testo
                            else:
                                testo = "Errore di comunicazione"
                                id_messaggio = testo

                            app_io_salva_messaggio(request.user, servizioScelto, cf, titolo, messaggio, dataScadenza, input_IUV, mezzo1, testobottone1, comandobottone1, mezzo2, testobottone2, comandobottone2, id_messaggio)
                            data.append((row[0].strip().upper(), correttezza_cf , risposta, testo))
                        contatore += 1

                salva_log(request.user,"Invio App IO massivo", "Invio n " + str(contatore) + " messaggi")
                request.session["multi_data"] = data  # <--- Salva i dati nella sessione
                return render(request, 'app_io_massivo.html', {'dati_catalogo': voci_catalogo, 'dati_servizio': voci_servizio, 'data': data, 'utente_abilitato': utente_abilitato })

            elif csv_file.name.endswith('.xlsx') or csv_file.name.endswith('.XLSX') or csv_file.name.endswith('.Xlsx'):
                wb = openpyxl.load_workbook(csv_file)
                sheet = wb.active
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    if row[0]:
                        cf = row[0].strip().upper()
                        titolo = row[1]
                        messaggio = row[2]
                        dataScadenza = ""
                        input_IUV = ""
                        mezzo1 = ""
                        testobottone1 = ""
                        comandobottone1 = ""
                        mezzo2 = ""
                        testobottone2 = ""
                        comandobottone2 = ""
                        if len(row) > 3 and row[3]:
                            dataScadenza = row[3]
                        if len(row) > 4 and row[4]:
                            input_IUV = row[4]
                        if len(row) > 5 and row[5]:
                            mezzo1 = row[5]+"1"
                        if len(row) > 6 and row[6]:
                            testobottone1 = row[6]
                        if len(row) > 7 and row[7]:
                            comandobottone1 = row[7]
                        if len(row) > 8 and row[8]:
                            mezzo2 = row[8]+"2"
                        if len(row) > 9 and row[9]:
                            testobottone2 = row[9]
                        if len(row) > 10 and row[10]:
                            comandobottone2 = row[10]

                        correttezza_cf = verifica_cf(cf)
                        if  correttezza_cf == 2: #CODICE FISCALE MINORENNE
                            data.append((row[0].strip().upper(), correttezza_cf , correttezza_cf, "CF di persona minorenne"))
                        elif correttezza_cf == -1: ##CODICE FISCALE SBAGLIATO
                            data.append((row[0].strip().upper(), correttezza_cf , correttezza_cf, "Codice fiscale errato"))
                        else:
                            risposta = app_io_verifica_utente_attivo(cf, servizioScelto.chiave_api)
                            if  risposta == 1: ##UTENTE ATTIVO
                                response = app_io_invio_messaggio(url, cf, servizioScelto.chiave_api, titolo, messaggio, dataScadenza, input_IUV, mezzo1, testobottone1, comandobottone1, mezzo2, testobottone2, comandobottone2)
                                id_messaggio = "Errore"
                                if response.status_code == 201:
                                    response_json = response.json()
                                    id_messaggio = response_json.get("id")
                                    testo = "ID: " + id_messaggio
                                    app_io_salva_messaggio(request.user, servizioScelto, cf, titolo, messaggio, dataScadenza, input_IUV, mezzo1, testobottone1, comandobottone1, mezzo2, testobottone2, comandobottone2, id_messaggio)
                                elif response.content.decode() == '91':
                                    testo = "Titolo troppo corto"
                                    risposta = 9
                                elif response.content.decode() == '92':
                                    testo = "Messaggio troppo corto"
                                    risposta = 9
                                else:
                                    testo = "Errore di comunicazione"
                            elif risposta == 0: ##UTENTE ATTIVO NON PERMESSO
                                testo = "Utente attivo ma servizio non attivo"
                            elif risposta == 7: ##UTENTE NON ATTIVO
                                testo = "Utente non attivo"
                            elif risposta == 8: ##UTENTE NON TROVATO
                                testo = "Utente non trovato"
                            else:
                                testo = "Errore di comunicazione"
                            data.append((row[0].strip().upper(), correttezza_cf , risposta, testo))
                        contatore += 1

                salva_log(request.user,"Invio App IO massivo", "Invio n " + str(contatore) + " messaggi")
                request.session["multi_data"] = data  # <--- Salva i dati nella sessione
                return render(request, 'app_io_massivo.html', {'dati_catalogo': voci_catalogo, 'dati_servizio': voci_servizio, 'data': data, 'utente_abilitato': utente_abilitato })

            else:
                salva_log(request.user,"Invio App IO massivo", "Errore caricamento file CSV")
                return render(request, 'app_io_massivo.html', {'dati_catalogo': voci_catalogo, 'dati_servizio': voci_servizio , 'error': 'Il file non è un CSV', 'utente_abilitato': utente_abilitato})
    else:
        utente_abilitato = False
    return render(request, 'app_io_massivo.html', { 'dati_catalogo': voci_catalogo, 'dati_servizio': voci_servizio, 'utente_abilitato': utente_abilitato})


def app_io_composer(request):
    voci_catalogo = AppIoCatalogoArgomenti.objects.all()
    voci_servizio = AppIoCatalogoServizi.objects.all()
    context = {}
    if request.user.id:
        utente_sessione = UtentiParametri.objects.get(id=request.user.id)
        utente_abilitato = utente_sessione.app_io_composer
        context = {
            'dati_catalogo': voci_catalogo,
            'dati_servizio': voci_servizio,
            'utente_abilitato': utente_abilitato,
            }
        if request.method == 'POST':
            if 'cf_csv' in request.FILES:
                csv_file = request.FILES['cf_csv']
                titoliVariabiliTesto = []
                if csv_file.name.endswith('.csv') or csv_file.name.endswith('.CSV') or csv_file.name.endswith('.Csv'):
                    csv_file_text = io.TextIOWrapper(csv_file.file, encoding='utf-8')
                    csv_reader = csv.reader(csv_file_text)
                    try:
                        prima_riga = next(csv_reader)
                    except StopIteration:
                         context.update({
                                'upload_csv': False,
                                'error_colonne': True,
                            })
                         return render(request, 'app_io_composer.html', context)

                    # controlla se la prima_riga ha i seguenti campi (CF,Scadenza,IUV,Mezzo1,TestoBottone1,Comando1,Mezzo2,TestoBottone2,Comando2)
                    expected_headers = ["CF", "SCADENZA", "IUV", "MEZZO1", "TESTOBOTTONE1", "COMANDO1", "MEZZO2", "TESTOBOTTONE2", "COMANDO2"]
                    
                    # Verifica headers
                    if len(prima_riga) < 9:
                         context.update({
                                'upload_csv': False,
                                'error_colonne': True,
                            })
                         return render(request, 'app_io_composer.html', context)
                    
                    for i, header in enumerate(expected_headers):
                        if prima_riga[i].strip().upper() != header:
                             context.update({
                                'upload_csv': False,
                                'error_colonne': True,
                            })
                             return render(request, 'app_io_composer.html', context)

                    ### VERIFICA TITOLI DOPPI
                    if len(prima_riga) > 9:
                        titoliVariabiliTesto = prima_riga[9:]
                        titoliVariabiliTesto = [elemento.upper() for elemento in titoliVariabiliTesto]
                        normalizzati = set()
                        for titolo in titoliVariabiliTesto:
                            titolo_lower = titolo.lower()
                            if titolo_lower in normalizzati:
                                context.update({
                                    'upload_csv': False,
                                    'error_colonne': True,
                                })
                                return render(request, 'app_io_composer.html', context)
                            normalizzati.add(titolo_lower)
                    
                    rows = list(csv_reader)
                    elencoMessaggi = []
                    
                    for row_idx, row in enumerate(rows, start=2): # start=2 perchè riga 1 è header
                        # Normalizza lunghezza riga
                        while len(row) < len(prima_riga):
                            row.append("")
                            
                        # Verifica IUV (col 3 -> indice 2)
                        if len(row) > 2 and row[2]:
                            iuv = "".join(row[2].split())
                            if len(iuv) != 18 or not iuv.isdigit(): # CHECK 16 NUMERI
                                context.update({
                                    'upload_csv': False,
                                    'error_validation_msg': f"Errore alla riga {row_idx}: IUV deve essere di 18 cifre numeriche.",
                                })
                                return render(request, 'app_io_composer.html', context)

                        # Verifica Mezzo1 (col 4 -> indice 3)
                        mezzo1 = row[3].strip().lower() if len(row) > 3 and row[3] else ""
                        testo1 = row[4].strip() if len(row) > 4 and row[4] else ""
                        comando1 = row[5].strip() if len(row) > 5 and row[5] else ""
                        
                        valid_mezzi = ["web", "mail", "sms", "tel"]
                        
                        if mezzo1:
                            if mezzo1 not in valid_mezzi:
                                context.update({
                                    'upload_csv': False,
                                    'error_validation_msg': f"Errore alla riga {row_idx}: Mezzo1 ha un valore non valido ({mezzo1}). Valori ammessi: web, mail, sms, tel.",
                                })
                                return render(request, 'app_io_composer.html', context)
                            if not testo1 or not comando1:
                                context.update({
                                    'upload_csv': False,
                                    'error_validation_msg': f"Errore alla riga {row_idx}: Se Mezzo1 è valorizzato, anche TestoBottone1 e Comando1 devono essere valorizzati.",
                                })
                                return render(request, 'app_io_composer.html', context)
                        else:
                            # Se Mezzo1 è vuoto, colonne 7, 8, 9 devono essere vuote (Mezzo2, Testo2, Comando2)
                            # Indici: Mezzo2=6, Testo2=7, Comando2=8
                            mezzo2 = row[6].strip() if len(row) > 6 and row[6] else ""
                            testo2 = row[7].strip() if len(row) > 7 and row[7] else ""
                            comando2 = row[8].strip() if len(row) > 8 and row[8] else ""
                            
                            if mezzo2 or testo2 or comando2:
                                context.update({
                                    'upload_csv': False,
                                    'error_validation_msg': f"Errore alla riga {row_idx}: Se Mezzo1 non è valorizzato, Mezzo2, TestoBottone2 e Comando2 devono essere vuoti.",
                                })
                                return render(request, 'app_io_composer.html', context)

                        # Verifica Mezzo2 (col 7 -> indice 6)
                        mezzo2 = row[6].strip().lower() if len(row) > 6 and row[6] else ""
                        testo2 = row[7].strip() if len(row) > 7 and row[7] else ""
                        comando2 = row[8].strip() if len(row) > 8 and row[8] else ""
                        
                        if mezzo2:
                            if mezzo2 not in valid_mezzi:
                                context.update({
                                    'upload_csv': False,
                                    'error_validation_msg': f"Errore alla riga {row_idx}: Mezzo2 ha un valore non valido ({mezzo2}). Valori ammessi: web, mail, sms, tel.",
                                })
                                return render(request, 'app_io_composer.html', context)
                            
                            if not testo2 or not comando2:
                                 context.update({
                                    'upload_csv': False,
                                    'error_validation_msg': f"Errore alla riga {row_idx}: Se Mezzo2 è valorizzato, anche TestoBottone2 e Comando2 devono essere valorizzati.",
                                })
                                 return render(request, 'app_io_composer.html', context)


                        ### verifica prima cella correttezza cf
                        if row[0]:
                            correttezza_cf = verifica_cf(row[0].strip().upper())
                            if correttezza_cf == 1:

                                row_modificata = list(row)  # Crea una copia della riga per modificarla
                                if len(row_modificata) > 3 and row_modificata[3]:
                                    row_modificata[3] = str(row_modificata[3]) + '1'
                                if len(row_modificata) > 6 and row_modificata[6]:
                                    row_modificata[6] = str(row_modificata[6]) + '2'
                                elencoMessaggi.append(row_modificata)


                    request.session["elencoMessaggi"] = elencoMessaggi
                    request.session["titoliVariabiliTesto"] = titoliVariabiliTesto

                elif csv_file.name.endswith('.xlsx') or csv_file.name.endswith('.XLSX') or csv_file.name.endswith('.Xlsx'):
                    wb = openpyxl.load_workbook(csv_file)
                    sheet = wb.active
                    prima_riga = [cell.value for cell in sheet[1]]
                    
                    # controlla se la prima_riga ha i seguenti campi (CF,Scadenza,IUV,Mezzo1,TestoBottone1,Comando1,Mezzo2,TestoBottone2,Comando2)
                    expected_headers = ["CF", "SCADENZA", "IUV", "MEZZO1", "TESTOBOTTONE1", "COMANDO1", "MEZZO2", "TESTOBOTTONE2", "COMANDO2"]
                    
                    if len(prima_riga) < 9:
                         context.update({
                                'upload_csv': False,
                                'error_colonne': True,
                            })
                         return render(request, 'app_io_composer.html', context)

                    for i, header in enumerate(expected_headers):
                        val = prima_riga[i]
                        if val is None: val = ""
                        if str(val).strip().upper() != header:
                             context.update({
                                'upload_csv': False,
                                'error_colonne': True,
                            })
                             return render(request, 'app_io_composer.html', context)

                    ### VERIFICA TITOLI DOPPI
                    if len(prima_riga) > 9:
                        titoliVariabiliTesto = prima_riga[9:]
                        titoliVariabiliTesto = [str(elemento).upper() for elemento in titoliVariabiliTesto if elemento is not None]
                        normalizzati = set()
                        for titolo in titoliVariabiliTesto:
                            titolo_lower = titolo.lower()
                            if titolo_lower in normalizzati:
                                context.update({
                                    'upload_csv': False,
                                    'error_colonne': True,
                                })
                                return render(request, 'app_io_composer.html', context)
                            normalizzati.add(titolo_lower)
                    
                    rows = list(sheet.iter_rows(min_row=2, values_only=True))
                    elencoMessaggi = []
                    
                    for row_idx, row in enumerate(rows, start=2):
                        # Convert tuple to list for mutability and padding
                        row_list = list(row)
                        while len(row_list) < len(prima_riga):
                            row_list.append(None)
                            
                        # Verifica IUV (col 3 -> indice 2)
                        if len(row_list) > 2 and row_list[2]:
                            iuv = "".join(str(row_list[2]).split())
                            if len(iuv) != 18 or not iuv.isdigit():
                                context.update({
                                    'upload_csv': False,
                                    'error_validation_msg': f"Errore alla riga {row_idx}: IUV deve essere di 18 cifre numeriche.",
                                })
                                return render(request, 'app_io_composer.html', context)
                                
                        # Verifica Mezzo1 (col 4 -> indice 3)
                        mezzo1 = str(row_list[3]).strip().lower() if len(row_list) > 3 and row_list[3] else ""
                        testo1 = str(row_list[4]).strip() if len(row_list) > 4 and row_list[4] else ""
                        comando1 = str(row_list[5]).strip() if len(row_list) > 5 and row_list[5] else ""
                        
                        valid_mezzi = ["web", "mail", "sms", "tel"]
                        
                        if mezzo1:
                            if mezzo1 not in valid_mezzi:
                                context.update({
                                    'upload_csv': False,
                                    'error_validation_msg': f"Errore alla riga {row_idx}: Mezzo1 ha un valore non valido ({mezzo1}). Valori ammessi: web, mail, sms, tel.",
                                })
                                return render(request, 'app_io_composer.html', context)
                            if not testo1 or not comando1:
                                context.update({
                                    'upload_csv': False,
                                    'error_validation_msg': f"Errore alla riga {row_idx}: Se Mezzo1 è valorizzato, anche TestoBottone1 e Comando1 devono essere valorizzati.",
                                })
                                return render(request, 'app_io_composer.html', context)
                        else:
                            # Se Mezzo1 è vuoto, colonne 7, 8, 9 devono essere vuote
                            mezzo2 = str(row_list[6]).strip() if len(row_list) > 6 and row_list[6] else ""
                            testo2 = str(row_list[7]).strip() if len(row_list) > 7 and row_list[7] else ""
                            comando2 = str(row_list[8]).strip() if len(row_list) > 8 and row_list[8] else ""
                            
                            if mezzo2 or testo2 or comando2:
                                context.update({
                                    'upload_csv': False,
                                    'error_validation_msg': f"Errore alla riga {row_idx}: Se Mezzo1 non è valorizzato, Mezzo2, TestoBottone2 e Comando2 devono essere vuoti.",
                                })
                                return render(request, 'app_io_composer.html', context)

                        # Verifica Mezzo2 (col 7 -> indice 6)
                        mezzo2 = str(row_list[6]).strip().lower() if len(row_list) > 6 and row_list[6] else ""
                        testo2 = str(row_list[7]).strip() if len(row_list) > 7 and row_list[7] else ""
                        comando2 = str(row_list[8]).strip() if len(row_list) > 8 and row_list[8] else ""
                        
                        if mezzo2:
                            if mezzo2 not in valid_mezzi:
                                context.update({
                                    'upload_csv': False,
                                    'error_validation_msg': f"Errore alla riga {row_idx}: Mezzo2 ha un valore non valido ({mezzo2}). Valori ammessi: web, mail, sms, tel.",
                                })
                                return render(request, 'app_io_composer.html', context)
                            
                            if not testo2 or not comando2:
                                 context.update({
                                    'upload_csv': False,
                                    'error_validation_msg': f"Errore alla riga {row_idx}: Se Mezzo2 è valorizzato, anche TestoBottone2 e Comando2 devono essere valorizzati.",
                                })
                                 return render(request, 'app_io_composer.html', context)


                        row_modificata = []
                        for idx, cell in enumerate(row_list):
                            if isinstance(cell, datetime.datetime):
                                formatted_date = cell.strftime('%d/%m/%Y')
                                row_modificata.append(formatted_date)
                            elif idx == 3 and cell is not None:
                                row_modificata.append(str(cell)+'1')
                            elif idx == 6 and cell is not None:
                                row_modificata.append(str(cell)+'2')
                            else:
                                row_modificata.append(cell if cell is not None else '')
                        elencoMessaggi.append(row_modificata)
                    request.session["elencoMessaggi"] = elencoMessaggi
                    request.session["titoliVariabiliTesto"] = titoliVariabiliTesto

                context.update({
                    'titoliVariabiliTesto': titoliVariabiliTesto,
                    'variabiliJS': json.dumps(titoliVariabiliTesto),
                    'upload_csv': True,
                })
                return render(request, 'app_io_composer.html', context)
    else:
        utente_abilitato = False
        context = {
            'dati_catalogo': voci_catalogo,
            'dati_servizio': voci_servizio,
            'utente_abilitato': utente_abilitato,
        }
    return render(request, 'app_io_composer.html', context)


def app_io_composer_conferma(request):
    voci_catalogo = AppIoCatalogoArgomenti.objects.all()
    voci_servizio = AppIoCatalogoServizi.objects.all()

    if request.user.id:
        context = {
            'dati_catalogo': voci_catalogo,
            'dati_servizio': voci_servizio,
        }
        utente_sessione = UtentiParametri.objects.get(id=request.user.id)
        utente_abilitato = utente_sessione.app_io_composer
        titoliVariabiliTesto = request.session.get("titoliVariabiliTesto", [])

        if not request.GET.get('page'):
            id_servizio = request.POST.get('sceltaServizio')
            titolo = request.POST.get('subject')
            md = MarkdownIt()
            testo_messaggio = md.render(request.POST.get('MessageArea'))
            servizioSceltoId = AppIoCatalogoServizi.objects.get(id=id_servizio)
            servizioScelto = servizioSceltoId.servizio
            request.session["servizioScelto"] = servizioScelto
            request.session["id_servizio"] = id_servizio
            request.session["titolo"] = titolo
            request.session["testo_messaggio"] = testo_messaggio
        else:
            servizioScelto = request.session.get("servizioScelto", [])
            id_servizio = request.session.get("id_servizio", [])
            titolo = request.session.get("titolo", [])
            testo_messaggio = request.session.get("testo_messaggio", [])

        elencoMessaggi = request.session.get("elencoMessaggi", [])
        paginator = Paginator(elencoMessaggi, 1)
        page_number = request.GET.get('page', 1)
        page_obj = paginator.get_page(page_number)

        titolo_personalizzato = titolo
        for i, variabile in enumerate(titoliVariabiliTesto):
            titolo_personalizzato = titolo_personalizzato.replace('[['+variabile+']]', str(elencoMessaggi[int(page_number)-1][i+9]))

        testo_personalizzato = testo_messaggio
        for i, variabile in enumerate(titoliVariabiliTesto):
            testo_personalizzato = testo_personalizzato.replace('[['+variabile+']]', str(elencoMessaggi[int(page_number)-1][i+9]))

        context.update({
            'utente_abilitato': utente_abilitato,
            'id_servizio': id_servizio,
            'servizioScelto': servizioScelto,
            'titolo': titolo,
            'titolo_personalizzato': titolo_personalizzato,
            'testo_messaggio': testo_messaggio,
            'testo_personalizzato': testo_personalizzato,
            "elencoMessaggi": elencoMessaggi,
            'page_obj': page_obj,
            })
        return render(request, 'app_io_composer_conferma_prev.html', context)
    else:
        utente_abilitato = False
    return render(request, 'app_io_composer_conferma_prev.html')


def app_io_composer_esito(request):
    app_io_parametri = AppIoParametri.objects.get(id=1)
    url = app_io_parametri.api_url + '/messages'

    if request.user.id:
        titoliVariabiliTesto = request.session.get("titoliVariabiliTesto", [])
        elencoMessaggi = request.session.get("elencoMessaggi", [])
        id_servizio = request.session.get("id_servizio", [])
        servizioSceltoObj = AppIoCatalogoServizi.objects.get(id=id_servizio)
        titolo = request.session.get("titolo", [])
        testo_messaggio = request.session.get("testo_messaggio", [])

        utente_sessione = UtentiParametri.objects.get(id=request.user.id)
        utente_abilitato = utente_sessione.app_io_composer

        data = []
        contatore = 0

        for row in elencoMessaggi:
            correttezza_cf = verifica_cf(row[0].strip().upper())
            if  correttezza_cf == 2: #CODICE FISCALE MINORENNE
                data.append((row[0].strip().upper(), correttezza_cf , correttezza_cf, "CF di persona minorenne"))
            elif correttezza_cf == -1: ##CODICE FISCALE SBAGLIATO
                data.append((row[0].strip().upper(), correttezza_cf , correttezza_cf, "Codice fiscale errato"))
            else:
                risposta = app_io_verifica_utente_attivo(row[0].strip().upper(), servizioSceltoObj.chiave_api)
                if  risposta == 1: ##UTENTE ATTIVO
                    titolo_personalizzato = titolo
                    for i, variabile in enumerate(titoliVariabiliTesto):
                        titolo_personalizzato = titolo_personalizzato.replace('[['+variabile+']]', str(row[i+9]))
                    testo_personalizzato = testo_messaggio
                    for i, variabile in enumerate(titoliVariabiliTesto):
                        testo_personalizzato = testo_personalizzato.replace('[['+variabile+']]', str(row[i+9]))
                    testo_personalizzato = html2text.html2text(testo_personalizzato)

                    response = app_io_invio_messaggio(url, row[0].strip().upper(), servizioSceltoObj.chiave_api, titolo_personalizzato, testo_personalizzato, row[1], row[2], row[3], row[4], row[5], row[6], row[7], row[8])
                    if response.status_code == 201:
                        response_json = response.json()
                        id_messaggio = response_json.get("id")
                        testo = "ID: " + id_messaggio
                    elif response.content.decode() == '91':
                        testo = "Titolo troppo corto"
                        id_messaggio = testo
                        risposta = 9
                    elif response.content.decode() == '92':
                        testo = "Messaggio troppo corto"
                        id_messaggio = testo
                        risposta = 9
                    else:
                        testo = "Errore di comunicazione"
                        id_messaggio = testo
                elif risposta == 0: ##UTENTE ATTIVO NON PERMESSO
                    testo = "Utente attivo ma servizio non attivo"
                    id_messaggio = testo
                elif risposta == 7: ##UTENTE NON ATTIVO
                    testo = "Utente non attivo"
                    id_messaggio = testo
                elif risposta == 8: ##UTENTE NON TROVATO
                    testo = "Utente non trovato"
                    id_messaggio = testo
                else:
                    testo = "Errore di comunicazione"
                    id_messaggio = testo
                app_io_salva_messaggio(request.user, servizioSceltoObj, row[0].strip().upper(), titolo_personalizzato, testo_personalizzato, row[1], row[2], row[3], row[4], row[5], row[6], row[7], row[8], id_messaggio)
                data.append((row[0].strip().upper(), correttezza_cf , risposta, testo))
            contatore += 1

    return render(request, 'app_io_composer_esito.html', {'data': data, 'utente_abilitato': utente_abilitato })


def app_io_verifica_utente(request):
    voci_catalogo = AppIoCatalogoArgomenti.objects.all()
    voci_servizio = AppIoCatalogoServizi.objects.all()
    if request.user.id:
        utente_sessione = UtentiParametri.objects.get(id=request.user.id)
        utente_abilitato = utente_sessione.app_io_verifica_singolo
        if request.method == 'POST':
            data = []
            cf = request.POST.get('input_CF')
            id_servizio = request.POST.get('sceltaServizio')
            servizioScelto = AppIoCatalogoServizi.objects.get(id=id_servizio)
            correttezza_cf = verifica_cf(cf)
            if  correttezza_cf == 2: #CODICE FISCALE MINORENNE
                testo = "CF di persona minorenne"
                risposta = 2
            elif correttezza_cf == -1: ##CODICE FISCALE SBAGLIATO
                testo = "Codice fiscale errato"
                risposta = -1
            else: ##CODICE FISCALE CORRETTO
                risposta = app_io_verifica_utente_attivo(cf, servizioScelto.chiave_api)
                if  risposta == 1: ##UTENTE ATTIVO
                    testo = "Utente attivo"
                elif risposta == 0: ##UTENTE ATTIVO NON PERMESSO
                    testo = "Utente attivo ma servizio non attivo"
                elif risposta == 7: ##UTENTE NON ATTIVO
                    testo = "Utente non attivo"
                elif risposta == 8: ##UTENTE NON TROVATO
                    testo = "Utente non trovato"
                else:
                    testo = "Errore di comunicazione"

            data.append(cf)
            data.append(servizioScelto.servizio)
            data.append(testo)
            data.append(risposta)
            salva_log(request.user,"Verifica utente App IO", "Verificato utente " + cf )
            return render(request, 'app_io_verifica_utente.html', {'dati_catalogo': voci_catalogo, 'dati_servizio': voci_servizio, 'data': data, 'utente_abilitato': utente_abilitato })
    else:
        utente_abilitato = False
    return render(request, 'app_io_verifica_utente.html', { 'dati_catalogo': voci_catalogo, 'dati_servizio': voci_servizio, 'utente_abilitato': utente_abilitato })


def app_io_verifica_massivo(request):
    voci_catalogo = AppIoCatalogoArgomenti.objects.all()
    voci_servizio = AppIoCatalogoServizi.objects.all()
    if request.user.id:
        utente_sessione = UtentiParametri.objects.get(id=request.user.id)
        utente_abilitato = utente_sessione.app_io_verifica_massivo
        if request.method == 'POST':
            csv_file = request.FILES['cf_csv']
            id_servizio = request.POST.get('sceltaServizio')
            servizioScelto = AppIoCatalogoServizi.objects.get(id=id_servizio)
            data = []
            contatore = 0
            if csv_file.name.endswith('.csv'):
                csv_file_text = io.TextIOWrapper(csv_file.file, encoding='utf-8')
                csv_reader = csv.reader(csv_file_text)
                for row in csv_reader:
                    if row[0]:  # Se row[0] non è vuoto o None
                        correttezza_cf = verifica_cf(row[0].strip().upper())
                        if correttezza_cf == 2: ##CODICE FISCALE MINORENNE
                            data.append(row[0].strip().upper() +" "+ "2")## + " Persona minorenne")
                        elif correttezza_cf == -1: ##CODICE FISCALE ERRATO
                            data.append(row[0].strip().upper() +" "+ "-1")## + " Codice fiscale errato")
                        else : ##CODICE FISCALE ok
                            data.append(row[0].strip().upper() + " " + str(app_io_verifica_utente_attivo(row[0].strip().upper(), servizioScelto.chiave_api)))## + " " + testo )
                        contatore += 1
                if request.user.id:
                    salva_log(request.user,"Verifica CF massivo", "Verificati n. " + str(contatore) + " CF")
                request.session["multi_data"] = data  # <--- Salva i dati nella sessione
                return render(request, 'app_io_verifica_massivo.html', {'dati_catalogo': voci_catalogo, 'dati_servizio': voci_servizio, 'data': data, 'servizioScelto': servizioScelto.servizio, 'utente_abilitato': utente_abilitato})

            elif csv_file.name.endswith('.xlsx') or csv_file.name.endswith('.XLSX') or csv_file.name.endswith('.Xlsx'):
                wb = openpyxl.load_workbook(csv_file)
                sheet = wb.active
                for row in sheet.iter_rows(min_row=1, values_only=True):
                    if row[0]:
                        correttezza_cf = verifica_cf(row[0].strip().upper())
                        if correttezza_cf == 2: ##CODICE FISCALE MINORENNE
                            data.append(row[0].strip().upper() +" "+ "2")## + " Persona minorenne")
                        elif correttezza_cf == -1: ##CODICE FISCALE ERRATO
                            data.append(row[0].strip().upper() +" "+ "-1")## + " Codice fiscale errato")
                        else : ##CODICE FISCALE ok
                            data.append(row[0].strip().upper() + " " + str(app_io_verifica_utente_attivo(row[0].strip().upper(), servizioScelto.chiave_api)))## + " " + testo )
                        contatore += 1
                if request.user.id:
                    salva_log(request.user,"Verifica CF massivo", "Verificati n. " + str(contatore) + " CF")
                request.session["multi_data"] = data  # <--- Salva i dati nella sessione
                return render(request, 'app_io_verifica_massivo.html', {'dati_catalogo': voci_catalogo, 'dati_servizio': voci_servizio, 'data': data, 'servizioScelto': servizioScelto.servizio, 'utente_abilitato': utente_abilitato})

            else:
                if request.user.id:
                    salva_log(request.user,"Verifica CF massivo", "Errore caricamento file CSV")
                return render(request, 'app_io_verifica_massivo.html', {'dati_catalogo': voci_catalogo, 'dati_servizio': voci_servizio , 'error': 'Il file non è un CSV', 'utente_abilitato': utente_abilitato })
    else:
        utente_abilitato = False
    return render(request, 'app_io_verifica_massivo.html', {'dati_catalogo': voci_catalogo, 'dati_servizio': voci_servizio, 'utente_abilitato': utente_abilitato })


def app_io_verifica_utente_attivo(cf, key_servizio):
    app_io_parametri = AppIoParametri.objects.get(id=1)
    url = app_io_parametri.api_url + '/profiles'
    headers = {
        'Content-Type': 'application/json',
        'Ocp-Apim-Subscription-Key': key_servizio
    }
    data = {
        'fiscal_code': cf
    }
    response = requests.post(url, headers=headers, json=data)
    if response.status_code == 200:
        response_json = response.json()
        sender_allowed = response_json.get("sender_allowed")
        if sender_allowed is True:
            return 1 #ok
        elif sender_allowed is False:
            return 0 #non permesso
        else:
            return 7 #utente non attivo
    elif response.status_code == 404:
        return 8 #utente non trovato
    else:
        return 9 #errore comunicazione


def app_io_storico_messaggi(request):
    voci_catalogo = AppIoCatalogoArgomenti.objects.all()
    voci_servizio = AppIoCatalogoServizi.objects.all()
    app_io_parametri = AppIoParametri.objects.get(id=1)

    # Otteniamo i parametri di ricerca dalla richiesta (GET)
    servizio = request.GET.get('sceltaServizio', '')
    cf = request.GET.get('cf', '')
    titolo = request.GET.get('titolo', '')
    data_da = request.GET.get('data_da', '')
    data_a = request.GET.get('data_a', '')

    messaggi = AppIoElencoMessaggi.objects.all().order_by('-id')

    if not request.user.is_superuser:
        messaggi = messaggi.filter(utente_id=request.user.id)

    if servizio:
        messaggi = messaggi.filter(servizio_id=servizio)
    if cf:
        messaggi = messaggi.filter(cf_destinatario__icontains=cf)
    if titolo:
        messaggi = messaggi.filter(titolo__icontains=titolo)


    if data_da:
        try:
            data_da = datetime.datetime.strptime(data_da, '%Y-%m-%d')  # Uso di datetime.strptime
        except ValueError:
            data_da = None

    if data_a:
        try:
            data_a = datetime.datetime.strptime(data_a, '%Y-%m-%d')
            data_a = data_a.replace(hour=23, minute=59, second=59)  # Aggiungo 23:59:59
        except ValueError:
            data_a = None

    if data_da and data_a:
        messaggi = messaggi.filter(timestamp__range=[data_da, data_a])
    elif data_da:
        messaggi = messaggi.filter(timestamp__gte=data_da)  # Maggiore o uguale a data_da
    elif data_a:
        messaggi = messaggi.filter(timestamp__lte=data_a)  # Minore o uguale a data_a

    paginator = Paginator(messaggi, 20)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)

    if request.user.id:
        utente_sessione = UtentiParametri.objects.get(id=request.user.id)
        utente_abilitato = utente_sessione.app_io_storico_messaggi

    else:
        utente_abilitato = False

    session_list = [
        {
            **{k: v for k, v in item.items() if k != 'timestamp'},'timestamp': item['timestamp'].strftime('%d/%m/%Y %H:%M'),
            # 'servizio_id': voci_servizio.get(id=item['servizio_id'], servizio=item['servizio_id']).servizio if item['servizio_id'] in voci_servizio.values_list('id', flat=True) else item['servizio_id'],
            # 'utente_id': item['utente_id'] if item['utente_id'] else 'Non specificato',
            #'esito': item['esito'] if item['esito'] else 'Non specificato',
        }

        for item in messaggi.values('id', 'servizio_id', 'titolo', 'cf_destinatario', 'timestamp', 'utente_id', 'esito')
    ]

    request.session['multi_messaggi'] = session_list


    return render(request, 'app_io_storico_messaggi.html', { 'dati_catalogo': voci_catalogo, 'dati_servizio': voci_servizio, 'utente_abilitato': utente_abilitato, 'messaggi': messaggi, 'page_obj': page_obj })


def app_io_salva_messaggio(utente, servizio_id, cf_destinatario, titolo, messaggio, scadenza, iuv,  mezzo1, testoBottone1, comando1, mezzo2, testoBottone2, comando2, esito):

    if scadenza != "" and scadenza is not None:
        dataScadenzaAux = normalizza_data(scadenza)
        if isinstance(dataScadenzaAux, datetime.datetime):
            oggi = datetime.datetime.today()
        else:
            oggi = datetime.date.today()
        if dataScadenzaAux < oggi:
            dataScadenzaAux = oggi + datetime.timedelta(days=365)

        timezone = 'Europe/Rome'
        if is_daylight_saving(dataScadenzaAux, timezone):
            dataScadenzaPulita = dataScadenzaAux.strftime('%Y-%m-%d')
        else:
            dataScadenzaPulita = dataScadenzaAux.strftime('%Y-%m-%d')

        scadenza = dataScadenzaPulita

    log_messaggio = AppIoElencoMessaggi(
        utente_id=utente,
        servizio_id=servizio_id,
        cf_destinatario=cf_destinatario,
        titolo=titolo,
        messaggio=messaggio,
        scadenza=scadenza if scadenza is not None else "None",
        iuv=iuv if iuv is not None else "None",
        mezzo1=mezzo1 if mezzo1 is not None else "None",
        testoBottone1=testoBottone1 if testoBottone1 is not None else "None",
        comando1=comando1 if comando1 is not None else "None",
        mezzo2=mezzo2 if mezzo2 is not None else "None",
        testoBottone2=testoBottone2 if testoBottone2 is not None else "None",
        comando2=comando2 if comando2 is not None else "None",
        esito=esito if esito is not None else "None"
        )
    log_messaggio.save()


def impostazioni_app_io_default_db():
    app_io_catalogo_argomenti= [
        ('Ambiente e animali'),
        ('Attività produttive e commercio'),
        ('Benessere sociale'),
        ('Casa e utenze'),
        ('Cultura, tempo libero e sport'),
        ('Educazione e formazione'),
        ('Lavori edilizi, catasto e urbanistica'),
        ('Mobilità e trasporti'),
        ('Redditi, patrimoni e fisco'),
        ('Servizi anagrafici e civici'),
        ('Servizi elettorali'),
        ('Suolo, spazi e beni pubblici'),
        ('Viaggi e turismo'),
        ('Vita lavorativa'),
        ('Sicurezza e protezione civile'),
        ('Giustizia e legge'),
        ('Fuori catagolo'),
    ]

    app_io_catalogo_servizi = [
        ('1', 'Animali domestici', 'C001001'),
        ('1', 'Discariche e isole ecologiche', 'C001002'),
        ('1', 'Disinfestazioni', 'C001003'),
        ('1', 'Energia e impianti', 'C001004'),
        ('1', 'Funghi, caccia e pesca', 'C001005'),
        ('1', 'Orti urbani', 'C001006'),
        ('1', 'Ritiro rifiuti ingombranti', 'C001007'),
        ('1', 'Verde pubblico', 'C001008'),
        ('2', 'Sportello unico per le attività produttive (SUAP)', 'C002001'),
        ('2', 'Sviluppo economico', 'C002002'),
        ('3', 'Alloggi sociali', 'C003001'),
        ('3', 'Assistenza domiciliare', 'C003002'),
        ('3', 'Servizi sociali', 'C003003'),
        ('4', 'Condomini e inquilini', 'C004001'),
        ('4', 'Diritti di rogito', 'C004002'),
        ('4', 'Imposta municipale unica (IMU)', 'C004003'),
        ('4', 'Raccolta differenziata dei rifiuti', 'C004004'),
        ('4', 'Servizi idrici', 'C004005'),
        ('4', 'Tassa sui rifiuti (TARI)', 'C004006'),
        ('5', 'Attività sportive', 'C005001'),
        ('5', 'Biblioteche', 'C005002'),
        ('5', 'Cinema e teatri', 'C005003'),
        ('5', 'Eventi e manifestazioni', 'C005004'),
        ('5', 'Musei civici', 'C005005'),
        ('6', 'Asilo nido', 'C006001'),
        ('6', 'Centri estivi e centri gioco', 'C006002'),
        ('6', 'Contributi allo studio', 'C006003'),
        ('6', 'Mensa scolastica', 'C006004'),
        ('6', 'Orientamento e formazione', 'C006005'),
        ('6', 'Scuola dell\'infanzia', 'C006006'),
        ('6', 'Scuola primaria e secondaria', 'C006007'),
        ('6', 'Trasporto scolastico', 'C006008'),
        ('7', 'Appalti pubblici', 'C008001'),
        ('7', 'Catasto', 'C008002'),
        ('7', 'Edilizia privata', 'C008003'),
        ('7', 'Lavori pubblici', 'C008004'),
        ('7', 'Passo carrabile', 'C008005'),
        ('8', 'Incidenti stradali', 'C009001'),
        ('8', 'Mobilità sostenibile', 'C009002'),
        ('8', 'Multe per violazione codice della strada', 'C009003'),
        ('8', 'Permessi per sosta e circolazione', 'C009004'),
        ('8', 'Rimozione veicoli', 'C009005'),
        ('8', 'Traffico', 'C009006'),
        ('9', 'Agevolazioni tributarie', 'C010001'),
        ('9', 'Depositi cauzionali', 'C010002'),
        ('9', 'Donazioni', 'C010003'),
        ('9', 'Riscossione coattiva e recupero crediti', 'C010004'),
        ('10', 'Accesso civico agli atti', 'C011001'),
        ('10', 'Anagrafe', 'C011002'),
        ('10', 'Carta d\'identità', 'C011003'),
        ('10', 'Cittadinanza italiana', 'C011004'),
        ('10', 'Consigli e Giunte comunali', 'C011005'),
        ('10', 'Diritti di segreteria', 'C011006'),
        ('10', 'Matrimonio, unioni civili e convivenze di fatto', 'C011007'),
        ('10', 'Nascita, adozione e riconoscimento', 'C011008'),
        ('10', 'Numeri civici', 'C011009'),
        ('10', 'Permesso di soggiorno', 'C011010'),
        ('10', 'Pesa pubblica', 'C011011'),
        ('10', 'Residenza', 'C011012'),
        ('10', 'Separazione e divorzio', 'C011013'),
        ('10', 'Servizi cimiteriali', 'C011014'),
        ('11', 'Elezioni', 'C012001'),
        ('11', 'Presidenti e scrutatori di seggio', 'C012002'),
        ('11', 'Tessera elettorale', 'C012003'),
        ('12', 'Alienazione di beni mobili e immobili', 'C014001'),
        ('12', 'Canone unico patrimoniale', 'C014002'),
        ('12', 'Legname', 'C014003'),
        ('12', 'Locazione locali e impianti comunali', 'C014004'),
        ('12', 'Patrimonio comunale', 'C014005'),
        ('12', 'Segnalazioni, suggerimenti e reclami', 'C014006'),
        ('13', 'Campeggi', 'C015001'),
        ('13', 'Imposta di soggiorno', 'C015002'),
        ('14', 'Bandi di concorso', 'C016001'),
        ('14', 'Cedolino per dipendenti', 'C016002'),
        ('14', 'Mensa e buoni pasto per dipendenti', 'C016003'),
        ('15', 'Avvocatura civica', 'C007001'),
        ('15', 'Protezione civile', 'C013001'),
        ('16', 'Giudici popolari', 'C007002'),
        ('16', 'Provvedimenti giudiziari', 'C007003'),
        ('16', 'Sanzioni amministrative', 'C007004')
    ]


    ###voci_catalogo = AppIoCatalogoArgomenti.objects.all()
    for i in range(1, len(app_io_catalogo_argomenti)+1):
        dato = AppIoCatalogoArgomenti(id=i, argomento=app_io_catalogo_argomenti[i-1] )
        dato.save()

    ###voci_servizio = AppIoCatalogoServizi.objects.all()
    for i in range(1, len(app_io_catalogo_servizi)+1):
        dato = AppIoCatalogoServizi(id=i, argomento_id=AppIoCatalogoArgomenti.objects.get(id=app_io_catalogo_servizi[i-1][0]),servizio=app_io_catalogo_servizi[i-1][1], codice_catalogo=app_io_catalogo_servizi[i-1][2] )
        dato.save()


def impostazioni_app_io(request):
    voci_parametro = AppIoParametri.objects.all()
    voci_catalogo = AppIoCatalogoArgomenti.objects.all()
    voci_servizio = AppIoCatalogoServizi.objects.all()
    if request.method == 'POST':
        if 'aggiona_servizi' in request.POST:
            impostazioni_app_io_recupera_servizi_in_selfcare()
        else:
            dato_parametro=AppIoParametri(1,request.POST.get('dati_url_api'),request.POST.get('dati_api_ket_master'))
            dato_parametro.save()
            ### counter_voci_servizio = AppIoCatalogoServizi.objects.count()
            ### for i in range(1, counter_voci_servizio):
            ###     if 'chiave[' + str(i) + ']' in request.POST:
            ###         chiave_api = request.POST.get('chiave[' + str(i) + ']')
            ###         chiave_da_aggiornare = AppIoCatalogoServizi.objects.get(id=i)
            ###         id_argomento = chiave_da_aggiornare.argomento_id
            ###         nome_servizio = chiave_da_aggiornare.servizio
            ###         dato = AppIoCatalogoServizi(i, id_argomento, nome_servizio, chiave_api )
            ###         dato.save()
            salva_log(request.user,"Impostazioni App IO", "modifica parametri")
    return render(request, 'impostazioni_app_io.html', { 'dati_parametro': voci_parametro, 'dati_catalogo': voci_catalogo, 'dati_servizio': voci_servizio })


def impostazioni_app_io_recupera_servizi_in_selfcare():
    app_io_parametri = AppIoParametri.objects.get(id=1)
    voce_fuori_catalogo = AppIoCatalogoArgomenti.objects.get(id=17)

    ##Cancello tutti i servizi con Fuori Catalogo
    AppIoCatalogoArgomenti.objects.all().delete()
    AppIoCatalogoServizi.objects.all().delete()
    impostazioni_app_io_default_db()

    counter_voci_servizio = AppIoCatalogoServizi.objects.count()
    url = app_io_parametri.api_url + '/manage/services'
    key_master = app_io_parametri.api_key_master
    headers = {
    'Content-Type': 'application/json',
    'Ocp-Apim-Subscription-Key': key_master
    }
    params = {
        'limit': '99'
        }

    response = requests.get(url, headers=headers, params=params)

    if response.status_code == 200:
        servizi_attivi = response.json()
        array_servizi_attivi = []
        ##creo array con servizio/id/stato
        for item in servizi_attivi["value"]:
            name = item.get("name")
            item_id = item.get("id")
            status_value = item.get("status", {}).get("value")
            if status_value == "approved" :
                array_servizi_attivi.append({
                    "name": name,
                    "id": item_id,
                    "status_value": status_value
                })

        ##verifico se ci sono servizi non in catalogo
        for i in range(0, len(array_servizi_attivi)):
            if not AppIoCatalogoServizi.objects.filter(servizio=array_servizi_attivi[i]["name"]):
                counter_voci_servizio = AppIoCatalogoServizi.objects.count()
                dato = AppIoCatalogoServizi(id=counter_voci_servizio+1, argomento_id=voce_fuori_catalogo, servizio=array_servizi_attivi[i]["name"], id_servizio=array_servizi_attivi[i]["id"], chiave_api = "")
                dato.save()

        ##aggiorno db con valori in array
        counter_voci_servizio = AppIoCatalogoServizi.objects.count()
        for i in range(1, counter_voci_servizio+1):
            chiave_da_aggiornare = AppIoCatalogoServizi.objects.get(id=i)
            for j in range(0, len(array_servizi_attivi)):
                if chiave_da_aggiornare.servizio == array_servizi_attivi[j]["name"]:
                    url = app_io_parametri.api_url + '/manage/services/' + array_servizi_attivi[j]["id"] + '/keys'
                    headers = {
                        'Content-Type': 'application/json',
                        'Ocp-Apim-Subscription-Key': key_master
                        }
                    response_key = requests.get(url, headers=headers)
                    if response_key.status_code == 200:
                        response_key_servizio = response_key.json()
                        dato = AppIoCatalogoServizi.objects.get(id=i)
                        dato.id_servizio=array_servizi_attivi[j]["id"]
                        dato.chiave_api = response_key_servizio["primary_key"]

                        dato.save()
        return 1


def impostazioni_app_io_2(request):
    voci_parametro = AppIoParametri.objects.all()
    voci_catalogo = AppIoCatalogoArgomenti.objects.all()
    voci_servizio = AppIoCatalogoServizi.objects.all()
    if request.method == 'POST':
        dato_parametro=AppIoParametri(1,request.POST.get('dati_url_api'),request.POST.get('dati_api_ket_master'))
        dato_parametro.save()
        for i in range(1, 200):
            if 'chiave[' + str(i) + ']' in request.POST:
                chiave_api = request.POST.get('chiave[' + str(i) + ']')
                chiave_da_aggiornare = AppIoCatalogoServizi.objects.get(id=i)
                id_argomento = chiave_da_aggiornare.argomento_id
                nome_servizio = chiave_da_aggiornare.servizio
                dato = AppIoCatalogoServizi(i, id_argomento, nome_servizio, chiave_api )
                dato.save()
        salva_log(request.user,"Impostazioni App IO", "modifica parametri")
    return render(request, 'impostazioni_app_io_2.html', { 'dati_parametro': voci_parametro, 'dati_catalogo': voci_catalogo, 'dati_servizio': voci_servizio })

