from django.shortcuts import render
from impostazioni.models import IpaParametri
from impostazioni.models import UtentiParametri

from .utils import salva_log
from .utils import converti_data
from .verifica_cf import verifica_cf_azienda
import requests
import json
import io
import csv
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment
from openpyxl.utils import get_column_letter

from django.http import HttpResponse


def ipa_export_excel(request):
    data = request.session.get("multi_data", [])  # Oppure recuperalo come preferisci
    wb = Workbook()
    ws = wb.active
    ws.append(["#", "CF", "Verifica", "Domicilio digitale"])

    # Dizionario per la larghezza massima di ogni colonna
    max_lengths = [len(str(cell.value)) for cell in ws[1]]  # Larghezze iniziali dall'intestazione

    if not data:
        return HttpResponse("Nessun dato disponibile per l'esportazione", status=400)

    for i, row in enumerate(data, 1):
        if isinstance(row, dict) and isinstance(row.get('data'), dict):
            # Caso in cui row è un dizionario singolo
            Ente = row['data']['denominazione']
            Verifica = "1"
            Domicilio = row['data'].get('pec', 'Domicilio digitale non trovato')

        elif isinstance(row, dict) and isinstance(row.get('data'), list):
            # Caso in cui row è un dizionario con una lista di dati
            Ente = row['data'][0]['denominazione']  # Prendi il valore di Ente dalla prima voce della lista
            Verifica = "1"
            pec_list = [sub_row.get('pec', '') for sub_row in row['data'] if sub_row.get('pec')]
            if pec_list:
                Domicilio = chr(10).join(pec_list)  # Usa newline per a capo nella stessa cella
                # Pulizia opzionale: rimuovi newline extra o spazi iniziali/fine per ogni PEC
                Domicilio = '\n'.join(pec.strip() for pec in Domicilio.split('\n') if pec.strip())
            else:
                Domicilio = "Domicilio digitale non trovato"

        else:
            Ente = str(row) if row else "Ente non trovato"
            Verifica = "-1"
            Domicilio = "Domicilio digitale non trovato"

        color = "FFFFFF"  # default bianco

        if Verifica == "1":
            color = "C6EFCE"  # verde chiaro
        else:
            color = "FFC7CE"  # rosso

        fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

        # Append senza normalizzare: preserva i newline in Domicilio
        ws.append([i, Ente, Verifica, Domicilio])

        # Applica fill a tutta la riga corrente
        current_row = ws.max_row
        for col in ws.iter_rows(min_row=current_row, max_row=current_row):
            for cell in col:
                cell.fill = fill

        # Calcola larghezze: per Domicilio, stima la larghezza orizzontale ignorando newline (solo per colonne)
        for j, cell in enumerate(ws[current_row]):
            value_str = str(cell.value) if cell.value is not None else ""
            # Per colonna D (indice 3), usa join con spazio per stimare larghezza max lineare
            if j == 3:  # Colonna Domicilio
                value_length = len(' '.join(value_str.split()))  # Ignora newline per larghezza orizzontale
            else:
                value_length = len(value_str)

            if len(max_lengths) <= j:
                max_lengths.append(value_length)
            else:
                max_lengths[j] = max(max_lengths[j], value_length)

    # Imposta larghezza colonne
    for col_idx, width in enumerate(max_lengths, 1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = min(width + 2, 50)  # Limita a 50 per non esagerare, +2 per padding

    # Allineamento wrap text per colonna D (Domicilio)
    wrap_alignment = Alignment(wrap_text=True, vertical='top', horizontal='left')  # Aggiunto horizontal='left' per chiarezza
    ws.column_dimensions['D'].alignment = wrap_alignment

    # Applica esplicitamente l'allineamento a tutte le celle della colonna D (per robustezza)
    for row in ws.iter_rows(min_row=2):  # Dal row 2 (dati, escludi intestazione)
        if len(row) >= 4:  # Assicura che ci sia la colonna D
            row[3].alignment = wrap_alignment  # Applica a cella D

    # Imposta altezze righe dinamicamente per gestire multipli newline in Domicilio
    for row_num in range(2, ws.max_row + 1):  # Dal row 2 in poi
        cell_d = ws.cell(row=row_num, column=4)  # Colonna D
        value_str = str(cell_d.value) if cell_d.value is not None else ""
        num_lines = len(value_str.split('\n'))  # Conta approssimativa delle righe (basata su newline)
        row_height = max(15, num_lines * 15)  # Altezza minima 15 per riga virtuale; max per multipli
        ws.row_dimensions[row_num].height = min(row_height, 200)  # Limita a 200 per non esagerare

    # Scrive l'Excel su un buffer
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=EsitoDomiciliDigitaliPA.xlsx'
    return response


def ipa_singola(request):
    if request.user.id:
        utente_sessione = UtentiParametri.objects.get(id=request.user.id)
        utente_abilitato = utente_sessione.ipa_singolo
        ipa_parametri = IpaParametri.objects.get(id=1)
        auth_id = ipa_parametri.auth_id
        codice_ipa = ""
        if request.method == 'POST':
            data = []
            occorrenze = ""
            testo_log = ""
            descrizione = request.POST.get('input_descrizione_ente')
            cf = request.POST.get('input_CF')
            codice_ipa = request.POST.get('input_codice_ente')
            if (descrizione != None): ##inserisco descrizione (anche più di una)
                testo_log = descrizione
                url = "https://www.indicepa.gov.it:443/ws/WS16DESAMMServices/api/WS16_DES_AMM"
                payload = {
                    "AUTH_ID": auth_id,
                    "DESCR": descrizione
                    }
                headers = {
                    'Content-Type': 'application/x-www-form-urlencoded'
                    }
                response = requests.post(url, data=payload, headers=headers)
                if response.status_code == 200:
                    content_str = response.content.decode('utf-8')
                    temp_data = json.loads(content_str)
                    occorrenze = temp_data['result']['num_items']
                    if occorrenze:
                        for indice in range(0 , occorrenze):
                            codice_ipa = temp_data['data'][indice]['cod_amm']
                            response = ipa_codice(auth_id, codice_ipa)
                            if response.status_code == 200:
                                risultato = response.json()
                            else:
                                risultato = {"error": f"Request failed with status code {response.status_code}"}
                            data.append(risultato)
                    else:
                        response = ipa_codice(auth_id, "999999")
                        risultato = response.json()
                        data.append(risultato)


            elif (cf != None): ##inserisco CF
                testo_log = cf
                verifica_cf_ipa = verifica_cf_azienda(cf)
                if verifica_cf_ipa == 1:
                    url = "https://www.indicepa.gov.it:443/ws/WS23DOMDIGCFServices/api/WS23_DOM_DIG_CF"
                    payload = {
                        "AUTH_ID": auth_id,
                        "CF": cf
                        }
                    headers = {
                        'Content-Type': 'application/x-www-form-urlencoded'
                        }
                    response = requests.post(url, data=payload, headers=headers)
                    if response.status_code == 200:
                        content_str = response.content.decode('utf-8')
                        temp_data = json.loads(content_str)
                        codice_ipa = temp_data['data'][0]['cod_amm']
                        response = ipa_codice(auth_id, codice_ipa)
                        if response.status_code == 200:
                            risultato = response.json()
                        else:
                            risultato = {"error": f"Request failed with status code {response.status_code}"}
                        data.append(risultato)
                else:
                    data.append("Codice fiscale non corretto")
            else: ##inserisco codice ipa
                response = ipa_codice(auth_id, codice_ipa)
                if response.status_code == 200:
                    risultato = response.json()
                else:
                    risultato = {"error": f"Request failed with status code {response.status_code}"}
                data.append(risultato)

            salva_log(request.user,"Verifica IndicePA", "Verificato domicilio ente " + testo_log )
            return render(request, 'ipa_singola.html', {'data': data, 'utente_abilitato': utente_abilitato })
    else:
        utente_abilitato = False
    return render(request, 'ipa_singola.html', { 'utente_abilitato': utente_abilitato })


def ipa_massiva(request):
    if request.user.id:
        utente_sessione = UtentiParametri.objects.get(id=request.user.id)
        utente_abilitato = utente_sessione.ipa_massivo
        ipa_parametri = IpaParametri.objects.get(id=1)
        auth_id = ipa_parametri.auth_id
        codice_ipa = ""
        if request.method == 'POST':
            csv_file = request.FILES['cf_csv']
            data = []
            contatore = 0
            url = "https://www.indicepa.gov.it:443/ws/WS16DESAMMServices/api/WS16_DES_AMM"
            headers = {
                'Content-Type': 'application/x-www-form-urlencoded'
                }
            if csv_file.name.endswith('.csv'):
                csv_file_text = io.TextIOWrapper(csv_file.file, encoding='utf-8')
                csv_reader = csv.reader(csv_file_text)

                for row in csv_reader:
                    if row[0]:  # Se row[0] non è vuoto o None
                        payload = {
                            "AUTH_ID": auth_id,
                            "DESCR": row[0].strip().upper()
                            }
                        response = requests.post(url, data=payload, headers=headers)
                        if response.status_code == 200:
                            content_str = response.content.decode('utf-8')
                            temp_data = json.loads(content_str)
                            occorrenze = temp_data['result']['num_items']
                            if (occorrenze == 0):
                                data.append(row[0].strip().upper())
                            for indice in range(0 , occorrenze):
                                codice_ipa = temp_data['data'][indice]['cod_amm']
                                response = ipa_codice(auth_id, codice_ipa)
                                if response.status_code == 200:
                                    risultato = response.json()
                                else:
                                    risultato = {"error": f"Request failed with status code {response.status_code}"}
                                data.append(risultato)
                        else:
                            data.append(row[0].strip().upper())
                        contatore += 1
                salva_log(request.user,"Verifica Ipa massivo", "Verificati n. " + str(contatore) + " CF")
                request.session["multi_data"] = data  # <--- Salva i dati nella sessione
                return render(request, 'ipa_massiva.html', {'data': data, 'utente_abilitato': utente_abilitato })
            elif csv_file.name.endswith('.xlsx') or csv_file.name.endswith('.XLSX') or csv_file.name.endswith('.Xlsx'):
                wb = openpyxl.load_workbook(csv_file)
                sheet = wb.active
                for row in sheet.iter_rows(min_row=1, values_only=True):
                    if row[0]:
                        row = list(row)
                        payload = {
                            "AUTH_ID": auth_id,
                            "DESCR": row[0].strip().upper()
                            }
                        response = requests.post(url, data=payload, headers=headers)
                        if response.status_code == 200:
                            content_str = response.content.decode('utf-8')
                            temp_data = json.loads(content_str)
                            occorrenze = temp_data['result']['num_items']
                            if (occorrenze == 0):
                                data.append(row[0].strip().upper())
                            for indice in range(0 , occorrenze):
                                codice_ipa = temp_data['data'][indice]['cod_amm']
                                response = ipa_codice(auth_id, codice_ipa)
                                if response.status_code == 200:
                                    risultato = response.json()
                                else:
                                    risultato = {"error": f"Request failed with status code {response.status_code}"}
                                data.append(risultato)
                        else:
                            data.append(row[0].strip().upper())
                        contatore += 1
                salva_log(request.user,"Verifica Ipa massivo", "Verificati n. " + str(contatore) + " CF")
                request.session["multi_data"] = data  # <--- Salva i dati nella sessione
                return render(request, 'ipa_massiva.html', {'data': data, 'utente_abilitato': utente_abilitato })
            else:
                salva_log(request.user,"Verifica Ipa massivo", "Errore caricamento file CSV")
                return render(request, 'ipa_massiva.html', {'error': 'Il file non è un CSV', 'utente_abilitato': utente_abilitato })
    else:
        utente_abilitato = False
    return render(request, 'ipa_massiva.html', { 'utente_abilitato': utente_abilitato })


def ipa_codice(auth_id, codice_ipa):
    url = "https://www.indicepa.gov.it:443/ws/WS20PECServices/api/WS20_PEC"
    payload = {
        "AUTH_ID": auth_id,
        "COD_AMM": codice_ipa
        }
    headers = {
        'Content-Type': 'application/x-www-form-urlencoded'
        }
    return requests.post(url, data=payload, headers=headers)


def impostazioni_ipa(request):
    parametri_ipa = IpaParametri.objects.all()
    if request.method == 'POST':
        auth_id = request.POST.get('auth_id')
        dati = IpaParametri(1, auth_id)
        dati.save()
        salva_log(request.user,"Impostazioni IPA", "modifica parametri")
    else:
        parametri_ipa = IpaParametri.objects.all()

    return render(request, 'impostazioni_ipa.html', {'parametri_ipa': parametri_ipa})

